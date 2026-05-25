def normalize_qty(value):
    try:
        if value is None or str(value).strip() == "":
            return 1
        return int(float(value))
    except Exception:
        return 1


def size_sort_key(size):
    s = str(size).strip()
    try:
        return (0, float(s))
    except Exception:
        return (1, s)


def merge_sizes(values):
    """
    旧兼容函数：仅合并尺码列表。
    """
    vals = [str(v).strip() for v in values if str(v).strip()]
    unique = []
    for v in vals:
        if v not in unique:
            unique.append(v)
    return " ".join(sorted(unique, key=size_sort_key))


def merge_size_quantity(group):
    """
    按商品数量展开尺码。
    例如：41码，数量3，输出为：41 41 41。
    """
    expanded = []

    for _, row in group.iterrows():
        size = str(row.get("尺码", "")).strip()
        if not size:
            continue

        qty = normalize_qty(row.get("数量", 1))
        for _ in range(max(qty, 0)):
            expanded.append(size)

    return " ".join(sorted(expanded, key=size_sort_key))


import os
import re
from datetime import datetime

import pandas as pd
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from order_secure_common import (
    get_output_dir, get_temp_dir, normalize_text, make_image_key, base64_to_image_file,
    safe_filename, col_letter_to_index, get_data_dir, ImageMatcher, load_image_map_for_categories,
)

IMAGE_WIDTH_PX = 140
IMAGE_HEIGHT_PX = 120
MATCH_TEXT_COLUMN = "图片匹配文本"


class RuleEngine:
    def __init__(self, rules):
        self.rules = []
        self.remove_words_by_category = {}

        for idx, rule in enumerate(rules or []):
            category = normalize_text(rule.get("category", ""))
            keyword = normalize_text(rule.get("keyword", ""))
            field = rule.get("field", "全部") or "全部"
            if not category or not keyword:
                continue

            self.rules.append({
                "index": idx,
                "category": category,
                "keyword": keyword,
                "field": field,
                "kw_len": len(keyword),
            })

            remove_words = str(rule.get("remove_words", "") or "")
            if remove_words:
                bucket = self.remove_words_by_category.setdefault(category, [])
                for w in remove_words.replace("，", ",").split(","):
                    w = normalize_text(w)
                    if w and w not in bucket:
                        bucket.append(w)

    def detect_category(self, short_name, title, raw_spec, remark=""):
        short_text = normalize_text(short_name)
        title_text = normalize_text(title)
        spec_text = normalize_text(raw_spec)
        remark_text = normalize_text(remark)

        best_score = -1
        best_category = "未分类"

        for rule in self.rules:
            kw = rule["keyword"]
            field = rule["field"]
            score = -1

            if field == "商品简称":
                if kw in short_text:
                    score = 1000 + rule["kw_len"]
            elif field == "销售规格":
                if kw in spec_text:
                    score = 700 + rule["kw_len"]
            elif field == "备注":
                if kw in remark_text:
                    score = 850 + rule["kw_len"]
            elif field == "货品标题":
                if kw in title_text:
                    score = 500 + rule["kw_len"]
            else:
                if kw in short_text:
                    score = 1000 + rule["kw_len"]
                elif kw in remark_text:
                    score = 850 + rule["kw_len"]
                elif kw in spec_text:
                    score = 700 + rule["kw_len"]
                elif kw in title_text:
                    score = 500 + rule["kw_len"]

            if score > best_score:
                best_score = score
                best_category = rule["category"]

        return best_category

    def clean_spec(self, raw_spec, category):
        spec = normalize_text(raw_spec)
        for word in self.remove_words_by_category.get(normalize_text(category), []):
            spec = spec.replace(word, "")
        return spec if spec else raw_spec


def split_items(value, sep):
    if value is None:
        return []
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return []
    return [x.strip() for x in text.split(sep) if x.strip()]


def split_spec_to_color_size(spec, split_char):
    spec = str(spec).strip()
    for sp in [split_char, "，", ",", " "]:
        if sp and sp in spec:
            parts = spec.rsplit(sp, 1)
            if len(parts) == 2:
                return parts[0].strip(), parts[1].strip()
    return spec, "未知"


def merge_text_values(values, max_chars=300):
    unique = []
    for value in values:
        text = str(value or "").strip()
        if not text or text.lower() == "nan":
            continue
        if text not in unique:
            unique.append(text)
    merged = "；".join(unique)
    return merged[:max_chars]


def get_template(system, template_name=""):
    templates = system.get("import_templates", [])
    template_name = (template_name or "").strip()

    if not templates:
        raise ValueError("当前整理系统未配置导入模板，请先打开“订单整理管理系统”新增模板")

    template = next((t for t in templates if t.get("name") == template_name), None)
    if not template:
        raise ValueError("请先选择一个有效的导入模板")

    return template


def read_by_template(file_path, template):
    df = pd.read_excel(file_path)
    mode = template.get("mode", "表头")
    item_sep = template.get("item_sep", ";") or ";"
    spec_split = template.get("spec_split", "，") or "，"
    rows = []

    if mode == "表头":
        short_col = template.get("short_name", "")
        spec_col = template.get("spec", "")
        qty_col = template.get("qty", "")
        remark_col = template.get("remark", "")
        if not {short_col, spec_col, qty_col}.issubset(set(df.columns)):
            raise ValueError(f"文件缺少模板指定表头：{os.path.basename(file_path)}")

        use_cols = [short_col, spec_col, qty_col]
        has_remark = bool(remark_col and remark_col in df.columns)
        if has_remark:
            use_cols.append(remark_col)

        temp = df[use_cols].copy()
        temp = temp.dropna(subset=[short_col, spec_col])

        for _, row in temp.iterrows():
            names = split_items(row[short_col], item_sep)
            specs = split_items(row[spec_col], item_sep)
            qtys = split_items(row[qty_col], item_sep)
            remarks = split_items(row[remark_col], item_sep) if has_remark else []
            max_len = max(len(names), len(specs), len(qtys), len(remarks), 1)

            for i in range(max_len):
                name = names[i] if i < len(names) else (names[-1] if names else "")
                spec = specs[i] if i < len(specs) else (specs[-1] if specs else "")
                qty = qtys[i] if i < len(qtys) else (qtys[-1] if qtys else "1")
                remark = remarks[i] if i < len(remarks) else (remarks[-1] if remarks else "")
                color, size = split_spec_to_color_size(spec, spec_split)
                rows.append({
                    "商品简称": name,
                    "货品标题": f"{name} 颜色: {color} 尺码: {size}",
                    "数量": qty,
                    "备注": remark,
                    "来源文件": os.path.basename(file_path),
                })
    else:
        title_idx = col_letter_to_index(template.get("title_col", "S"))
        qty_idx = col_letter_to_index(template.get("qty_col", "V"))
        remark_col = template.get("remark_col", "")
        remark_idx = col_letter_to_index(remark_col) if remark_col else None
        required_idx = [title_idx, qty_idx] + ([remark_idx] if remark_idx is not None else [])
        if df.shape[1] <= max(required_idx):
            raise ValueError(f"文件列数不足：{os.path.basename(file_path)}")

        temp = df.iloc[:, required_idx].copy()
        temp.columns = ["货品标题", "数量"] + (["备注"] if remark_idx is not None else [])
        temp = temp.dropna(subset=["货品标题"])

        def guess_short_name(title):
            m = re.search(r"^(.*?)\s*颜色[:：]", str(title))
            return m.group(1).strip() if m else str(title).strip()

        for _, row in temp.iterrows():
            rows.append({
                "商品简称": guess_short_name(row["货品标题"]),
                "货品标题": row["货品标题"],
                "数量": row["数量"],
                "备注": row.get("备注", ""),
                "来源文件": os.path.basename(file_path),
            })

    out = pd.DataFrame(rows)
    if "备注" not in out.columns:
        out["备注"] = ""
    out["数量"] = pd.to_numeric(out["数量"], errors="coerce").fillna(1).astype(int)
    return out


def extract_size(title):
    m = re.search(r"尺码[:：]\s*([0-9.]+)", str(title))
    return m.group(1) if m else "未知"


def extract_raw_spec(title):
    m = re.search(r"颜色[:：]\s*(.*?)\s*尺码", str(title))
    return normalize_text(m.group(1)) if m else "未知"


def size_sort_key(size):
    try:
        return float(size)
    except Exception:
        return 999


def merge_sizes(series):
    sizes = [str(x).strip() for x in series if str(x).strip() and str(x).strip() != "未知"]
    return " ".join(sorted(sizes, key=size_sort_key))


def get_stall(category, stall_map):
    return stall_map.get(normalize_text(category), "未设置档口")


def safe_sheet_name(name):
    name = re.sub(r'[\\/:*?\[\]]', "_", str(name).strip())
    return name[:31] if name else "未设置档口"


def prepare_image_for_excel(image_b64, category, spec):
    temp_dir = get_temp_dir()
    raw = os.path.join(temp_dir, f"{safe_filename(category)}_{safe_filename(spec)}_raw")
    png = os.path.join(temp_dir, f"{safe_filename(category)}_{safe_filename(spec)}.png")
    base64_to_image_file(image_b64, raw)
    return prepare_image_file_for_excel(raw, category, spec)


def prepare_image_file_for_excel(source_path, category, spec):
    temp_dir = get_temp_dir()
    png = os.path.join(temp_dir, f"{safe_filename(category)}_{safe_filename(spec)}.png")

    img = PILImage.open(source_path)
    if img.mode in ("RGBA", "LA"):
        bg = PILImage.new("RGBA", img.size, (255, 255, 255, 255))
        bg.paste(img, mask=img.getchannel("A"))
        img = bg.convert("RGB")
    else:
        img = img.convert("RGB")

    img.thumbnail((IMAGE_WIDTH_PX, IMAGE_HEIGHT_PX), PILImage.LANCZOS)
    canvas = PILImage.new("RGB", (IMAGE_WIDTH_PX, IMAGE_HEIGHT_PX), (255, 255, 255))
    canvas.paste(img, ((IMAGE_WIDTH_PX - img.width) // 2, (IMAGE_HEIGHT_PX - img.height) // 2))
    canvas.save(png)
    return png


def prepare_image_item_for_excel(item, category, spec, image_cache=None):
    if not item:
        return ""

    cache_key = make_image_key(category, spec)
    if image_cache is not None and cache_key in image_cache:
        return image_cache[cache_key]

    prepared = ""
    image_file = item.get("image_file") or item.get("file") or ""
    if image_file:
        source = image_file
        if not os.path.isabs(source):
            source = os.path.join(get_data_dir(), image_file)
        if os.path.exists(source):
            prepared = prepare_image_file_for_excel(source, category, spec)

    image_b64 = item.get("image_base64")
    if not prepared and image_b64:
        prepared = prepare_image_for_excel(image_b64, category, spec)

    if image_cache is not None:
        image_cache[cache_key] = prepared

    return prepared


def style_and_images(output_file, image_map, image_cache=None):
    matcher = ImageMatcher(image_map)
    wb = load_workbook(output_file)
    for ws in wb.worksheets:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = {normalize_text(ws.cell(row=1, column=col).value): col for col in range(1, ws.max_column + 1)}
        category_col = headers.get("分类", 1)
        spec_col = headers.get("规格", 2)
        image_col = headers.get("图片", 3)
        remark_col = headers.get("备注")
        match_col = headers.get(MATCH_TEXT_COLUMN)
        image_col_letter = ws.cell(row=1, column=image_col).column_letter
        if match_col:
            ws.column_dimensions[ws.cell(row=1, column=match_col).column_letter].hidden = True

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for col in range(1, ws.max_column + 1):
            col_letter = ws.cell(row=1, column=col).column_letter
            header = normalize_text(ws.cell(row=1, column=col).value)
            if header == "图片":
                ws.column_dimensions[col_letter].width = 20
            elif header == "尺码":
                ws.column_dimensions[col_letter].width = 42
            elif header == "规格":
                ws.column_dimensions[col_letter].width = 28
            else:
                ws.column_dimensions[col_letter].width = 16

        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = IMAGE_HEIGHT_PX * 0.75
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            category = normalize_text(ws.cell(row=row, column=category_col).value)
            spec = normalize_text(ws.cell(row=row, column=spec_col).value)
            remark = ws.cell(row=row, column=remark_col).value if remark_col else ""
            match_text = ws.cell(row=row, column=match_col).value if match_col else ""
            item = matcher.find(category, spec, remark, match_text)

            if item:
                try:
                    png_path = prepare_image_item_for_excel(item, category, spec, image_cache)
                    if png_path:
                        img = XLImage(png_path)
                        img.width = IMAGE_WIDTH_PX
                        img.height = IMAGE_HEIGHT_PX
                        ws.add_image(img, f"{image_col_letter}{row}")
                        ws.cell(row=row, column=image_col).value = ""
                except Exception:
                    ws.cell(row=row, column=image_col).value = ""
    wb.save(output_file)


def build_result(files, system, template_name=""):
    rules = system.get("category_rules", [])
    if not rules:
        raise ValueError("当前整理系统未配置分类规则，请先在后端数据管理中配置")

    template = get_template(system, template_name)
    rule_engine = RuleEngine(rules)

    df = pd.concat([read_by_template(f, template) for f in files], ignore_index=True)
    df["原始规格"] = df["货品标题"].apply(extract_raw_spec)
    df["尺码"] = df["货品标题"].apply(extract_size)
    if "备注" not in df.columns:
        df["备注"] = ""

    df["分类"] = df.apply(lambda r: rule_engine.detect_category(r.get("商品简称", ""), r["货品标题"], r["原始规格"], r.get("备注", "")), axis=1)
    df["规格"] = df.apply(lambda r: rule_engine.clean_spec(r["原始规格"], r["分类"]), axis=1)
    df["档口"] = df["分类"].apply(lambda c: get_stall(c, system.get("stall_map", {})))
    df[MATCH_TEXT_COLUMN] = df.apply(
        lambda r: merge_text_values([r.get("规格", ""), r.get("原始规格", ""), r.get("备注", ""), r.get("货品标题", ""), r.get("商品简称", "")]),
        axis=1
    )

    return (
        df.groupby(["档口", "分类", "规格"], dropna=False)
        .apply(lambda g: pd.Series({
            "尺码": merge_size_quantity(g),
            "数量": sum(normalize_qty(v) for v in g["数量"]),
            "备注": merge_text_values(g["备注"]),
            MATCH_TEXT_COLUMN: merge_text_values(g[MATCH_TEXT_COLUMN], max_chars=600),
        }))
        .reset_index()
        .sort_values(by=["档口", "分类", "规格"])
    )


def generate_order_file(files, system, output_mode="按档口分Sheet", template_name=""):
    result = build_result(files, system, template_name)
    used_categories = sorted({normalize_text(c) for c in result["分类"].dropna().tolist() if normalize_text(c)})
    image_map = load_image_map_for_categories(system, used_categories)
    image_cache = {}
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = get_output_dir()

    if output_mode == "合并一个Sheet":
        out_file = os.path.join(out_dir, f"订单整理文档_合并_{ts}.xlsx")
        out_df = result[["档口", "分类", "规格", "尺码", "数量", "备注", MATCH_TEXT_COLUMN]].copy()
        out_df.insert(3, "图片", "")
        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            out_df.to_excel(writer, sheet_name="全部订单", index=False)
        style_and_images(out_file, image_map, image_cache)
        return out_file

    if output_mode == "按档口分文档":
        batch = os.path.join(out_dir, f"订单整理文档_按档口分文档_{ts}")
        os.makedirs(batch, exist_ok=True)
        for stall, stall_df in result.groupby("档口", dropna=False):
            out_file = os.path.join(batch, f"{safe_sheet_name(stall)}_{ts}.xlsx")
            out_df = stall_df[["分类", "规格", "尺码", "数量", "备注", MATCH_TEXT_COLUMN]].copy()
            out_df.insert(2, "图片", "")
            with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
                out_df.to_excel(writer, sheet_name=safe_sheet_name(stall), index=False)
            style_and_images(out_file, image_map, image_cache)
        return batch

    out_file = os.path.join(out_dir, f"订单整理文档_分Sheet_{ts}.xlsx")
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        for stall, stall_df in result.groupby("档口", dropna=False):
            out_df = stall_df[["分类", "规格", "尺码", "数量", "备注", MATCH_TEXT_COLUMN]].copy()
            out_df.insert(2, "图片", "")
            out_df.to_excel(writer, sheet_name=safe_sheet_name(stall), index=False)
    style_and_images(out_file, image_map, image_cache)
    return out_file
