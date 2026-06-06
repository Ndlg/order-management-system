def normalize_qty(value):
    try:
        if value is None or str(value).strip() == "":
            return 1
        return int(float(value))
    except Exception:
        return 1


def size_sort_key(size):
    s = str(size).strip()
    try:
        return (0, float(s))
    except Exception:
        return (1, s)


def merge_size_quantity(group):
    """
    按商品数量展开尺码。
    例如：41码，数量3，输出为：41 41 41。
    """
    expanded = []

    for _, row in group.iterrows():
        size = str(row.get("尺码", "")).strip()
        if not size:
            continue

        qty = normalize_qty(row.get("数量", 1))
        for _ in range(max(qty, 0)):
            expanded.append(size)

    return " ".join(sorted(expanded, key=size_sort_key))


import os
import re
from datetime import datetime

import pandas as pd
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from utils.order_secure_common import (
    get_output_dir, get_temp_dir, normalize_text, make_image_key, base64_to_image_file,
    safe_filename, col_letter_to_index, get_data_dir, ImageMatcher, load_image_map_for_categories,
    normalize_image_match_text, normalize_match_text,
)
from core.five_field_normalizer import (
    SHOE_FIELD,
    SPEC_FIELD,
    SIZE_FIELD,
    QUANTITY_FIELD,
    REMARK_FIELD,
    RAW_SHOE_FIELD,
    RAW_SPEC_FIELD,
    make_five_field_item,
)
from core.shoe_rule_engine import detect_category_from_rules, detect_output_shoe_from_rules
from core.waybill_raw_contract import RAW_WAYBILL_TEMPLATE_NAME as WAYBILL_RAW_TEMPLATE_NAME
from core.waybill_raw_contract import RAW_WAYBILL_TEXT_COLUMN as WAYBILL_RAW_TEXT_COLUMN
from core.waybill_raw_contract import PROCESSED_WAYBILL_TEMPLATE_NAME
from core.waybill_raw_contract import RAW_WAYBILL_MODE
from core.waybill_raw_pipeline import is_waybill_raw_template, parse_raw_waybill_dataframe

IMAGE_WIDTH_PX = 140
IMAGE_HEIGHT_PX = 120
MATCH_TEXT_COLUMN = "图片匹配文本"
SHOE_CATEGORY_COLUMN = "鞋款分类"
RAW_SHORT_NAME_COLUMN = "原始商品简称"
WEAK_SPEC_VALUES = {"", "未知", "无", "无规格", "默认", "默认规格", "均码", "拍下备注", "看备注"}
REMARK_SPEC_LABELS = ("颜色分类", "颜色", "规格", "款式", "鞋款", "sku", "货号")
REMARK_SIZE_LABELS = ("鞋码", "尺码", "码数", "size")
REMARK_HEADER_KEYWORDS = (
    "卖家备注",
    "商家备注",
    "订单备注",
    "买家备注",
    "买家留言",
    "客户备注",
    "备注",
    "留言",
)
SIZE_TEXT_RE = re.compile(
    r"(?i)(?<!\d)(?:尺码|码数|鞋码|size)?\s*[:：=]?\s*((?:3[0-9]|4[0-9]|5[0-2])(?:\.5)?)\s*(?:码|m|M)?(?!\d)"
)
SHOP_CODE_RE = re.compile(r"(?i)(?:秒|范)\s*\d+[a-z]*")


def strip_shop_codes(value):
    text = normalize_text(value)
    if not text:
        return ""
    cleaned = SHOP_CODE_RE.sub(" ", text)
    return re.sub(r"[\s,，;；/|_\-]+", " ", cleaned).strip()


def clean_short_name_for_output(value, category=""):
    """
    商品简称里经常混入店铺/供货代号，例如“4.0 秒31”“范14 one帆布kw”。
    这些代号可以参与旧规则兼容识别，但不能作为最终归类键，否则同鞋款会被拆开。
    """
    text = normalize_text(value)
    cleaned = strip_shop_codes(text)
    if cleaned and not (re.fullmatch(r"\d+(?:\.\d+)?", cleaned) and cleaned not in {"4.0", "5.0"}):
        return cleaned
    if text:
        return text
    return normalize_text(category) or text


class RuleEngine:
    def __init__(self, rules):
        self.rules = []
        self.remove_words_by_category = {}

        for idx, rule in enumerate(rules or []):
            category = normalize_text(rule.get("category", ""))
            keyword = normalize_text(rule.get("keyword", ""))
            field = rule.get("field", "全部") or "全部"
            if not category or not keyword:
                continue

            self.rules.append({
                "index": idx,
                "category": category,
                "keyword": keyword,
                "field": field,
                "kw_len": len(keyword),
                "output_shoe": normalize_text(rule.get("output_shoe") or rule.get("shoe_name") or ""),
            })

            remove_words = str(rule.get("remove_words", "") or "")
            if remove_words:
                bucket = self.remove_words_by_category.setdefault(category, [])
                for w in remove_words.replace("，", ",").split(","):
                    w = normalize_text(w)
                    if w and w not in bucket:
                        bucket.append(w)

    def detect_category(self, short_name, title, raw_spec, remark="", size="", qty=""):
        """
        V7.8.0: 分类语义调整为“鞋款分类”。
        规则优先消费统一五字段，尤其是鞋款；货品标题只保留给旧规则兼容。
        """
        return detect_category_from_rules(self.rules, short_name, title, raw_spec, remark, size, qty)

    def detect_output_shoe(self, short_name, title, raw_spec, remark="", size="", qty="", category=""):
        return detect_output_shoe_from_rules(
            self.rules,
            short_name=short_name,
            title=title,
            raw_spec=raw_spec,
            remark=remark,
            size=size,
            qty=qty,
            category=category,
        )

    def clean_spec(self, raw_spec, category):
        spec = normalize_text(raw_spec)
        for word in self.remove_words_by_category.get(normalize_text(category), []):
            spec = spec.replace(word, "")
        return spec if spec else raw_spec


def split_items(value, sep):
    if value is None:
        return []
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return []
    separators = [";", "；", "\r\n", "\n"]
    if sep and sep not in separators:
        separators.append(sep)
    pattern = "|".join(re.escape(x) for x in separators if x)
    return [x.strip() for x in re.split(pattern, text) if x.strip()]


def is_weak_spec_text(value):
    return normalize_text(value) in WEAK_SPEC_VALUES


def is_abnormal_product_spec(value):
    text = normalize_text(value)
    if is_weak_spec_text(text):
        return True
    if is_valid_order_size(text):
        return True
    if re.fullmatch(r"\d+(?:\.0)?", text):
        try:
            number = int(float(text))
            return 1 <= number <= 20
        except Exception:
            return True
    return False


def split_spec_to_color_size(spec, split_char):
    spec = str(spec).strip()
    for sp in [split_char, "，", ",", " "]:
        if sp and sp in spec:
            parts = spec.rsplit(sp, 1)
            if len(parts) == 2:
                return parts[0].strip(), parts[1].strip()
    return spec, "未知"


def split_remark_for_items(remark_text, item_sep, item_count):
    fragments = split_items(remark_text, item_sep)
    if not fragments:
        return []

    def is_spec_fragment(value):
        label = normalize_text(str(value).split(":", 1)[0].split("：", 1)[0])
        return any(k in label.lower() for k in REMARK_SPEC_LABELS)

    def is_size_fragment(value):
        label = normalize_text(str(value).split(":", 1)[0].split("：", 1)[0])
        return any(k in label.lower() for k in REMARK_SIZE_LABELS)

    groups = []
    current = []
    seen_spec = False
    for fragment in fragments:
        if current and is_spec_fragment(fragment) and (seen_spec or item_count <= 1):
            groups.append(current)
            current = []
            seen_spec = False
        current.append(fragment)
        if is_spec_fragment(fragment):
            seen_spec = True
    if current:
        groups.append(current)

    structured = [
        merge_text_values(group)
        for group in groups
        if any(is_spec_fragment(x) or is_size_fragment(x) for x in group)
    ]
    if len(structured) > 1:
        return structured
    if len(groups) == item_count:
        return [merge_text_values(group) for group in groups]
    if len(fragments) == item_count:
        return fragments
    return [merge_text_values(fragments)]


def remark_label_value(fragment):
    text = str(fragment or "").strip()
    if not text or text.lower() == "nan":
        return "", ""
    parts = re.split(r"[:：=]", text, maxsplit=1)
    if len(parts) != 2:
        return "", ""
    return normalize_text(parts[0]).lower(), parts[1].strip()


def extract_remark_spec_size(remark_text, item_sep=";"):
    spec = ""
    size = ""
    for fragment in split_items(remark_text, item_sep):
        label, value = remark_label_value(fragment)
        if not label or not value:
            continue
        if any(k in label for k in REMARK_SPEC_LABELS):
            spec = normalize_text(value)
        if any(k in label for k in REMARK_SIZE_LABELS):
            found = extract_size_from_text(value)
            size = found or normalize_text(value)
    if not size:
        size = extract_size_from_text(remark_text)
    return spec, size


def merge_text_values(values, max_chars=300):
    unique = []
    for value in values:
        text = str(value or "").strip()
        if not text or text.lower() == "nan":
            continue
        if text not in unique:
            unique.append(text)
    merged = "；".join(unique)
    return merged[:max_chars]


def find_order_remark_columns(columns, preferred=""):
    headers = [str(c).strip() for c in columns if str(c).strip()]
    selected = []

    preferred = str(preferred or "").strip()
    if preferred and preferred in headers:
        selected.append(preferred)

    for keyword in REMARK_HEADER_KEYWORDS:
        for header in headers:
            if header in selected:
                continue
            if keyword and keyword in header:
                selected.append(header)

    return selected


def merge_order_remark(row, remark_cols):
    return merge_text_values(row.get(col, "") for col in remark_cols)


def extract_size_from_text(value):
    text = str(value or "")
    matches = SIZE_TEXT_RE.findall(text)
    return matches[-1].strip() if matches else ""


def is_valid_order_size(value):
    return bool(re.fullmatch(r"(?:3[0-9]|4[0-9]|5[0-2])(?:\.5)?", normalize_text(value)))


def normalize_order_size_text(value):
    text = normalize_text(value)
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def resolve_effective_size(current_size, *extra_texts):
    size = normalize_order_size_text(current_size)
    if size and size not in {"未知", "默认", "nan"} and is_valid_order_size(size):
        return size
    for text in extra_texts:
        found = extract_size_from_text(text)
        if found:
            return found
    return current_size


def resolve_five_field_spec_size(spec, split_char, remark, item_sep, size_value=""):
    if normalize_text(size_value):
        color = str(spec).strip()
        size = normalize_order_size_text(size_value)
    else:
        color, size = split_spec_to_color_size(spec, split_char)
        size = normalize_order_size_text(size)
    remark_spec, remark_size = extract_remark_spec_size(remark, item_sep)
    if remark_spec and is_abnormal_product_spec(color):
        color = remark_spec
    if remark_size and (is_weak_spec_text(size) or not is_valid_order_size(size)):
        size = remark_size
    return color, size


def read_waybill_raw_template(df, file_path, template, rule_config=None):
    rows = []
    for parsed in parse_raw_waybill_dataframe(df, file_path, template, rule_config):
        if parsed.get("解析状态") != "已解析" or not parsed.get("尺码"):
            continue
        raw_text = parsed.get("原始打印信息", "")
        rows.append(make_five_field_item(
            shoe=parsed.get("商品简称", ""),
            spec=parsed.get("规格", ""),
            size=parsed.get("尺码", ""),
            quantity=parsed.get("数量", 1),
            remark=parsed.get("备注", ""),
            source_file=os.path.basename(file_path),
            raw_shoe=parsed.get("商品简称", ""),
            raw_spec=parsed.get("规格", ""),
            raw_text=raw_text,
        ).to_order_row())

    out = pd.DataFrame(rows)
    if "备注" not in out.columns:
        out["备注"] = ""
    if "数量" not in out.columns:
        out["数量"] = 1
    out["数量"] = pd.to_numeric(out["数量"], errors="coerce").fillna(1).astype(int)
    return out


def get_template(system, template_name=""):
    templates = system.get("import_templates", [])
    template_name = (template_name or "").strip()

    if not templates:
        raise ValueError("当前整理系统未配置导入模板，请先打开“订单整理管理系统”新增模板")

    template = next((t for t in templates if t.get("name") == template_name), None)
    if not template:
        raise ValueError("请先选择一个有效的导入模板")

    return template


def read_by_template(file_path, template, rule_config=None):
    df = pd.read_excel(file_path)
    mode = template.get("mode", "表头")
    item_sep = template.get("item_sep", ";") or ";"
    spec_split = template.get("spec_split", "，") or "，"
    rows = []

    if is_waybill_raw_template(template, df.columns):
        return read_waybill_raw_template(df, file_path, template, rule_config)

    if mode == "表头":
        short_col = template.get("short_name", "")
        spec_col = template.get("spec", "")
        size_col = template.get("size", "")
        qty_col = template.get("qty", "")
        remark_col = template.get("remark", "")
        raw_text_col = WAYBILL_RAW_TEXT_COLUMN if WAYBILL_RAW_TEXT_COLUMN in df.columns else ""
        if template.get("name") == PROCESSED_WAYBILL_TEMPLATE_NAME and "原始打印信息" in df.columns:
            raw_text_col = "原始打印信息"
        if not {short_col, spec_col, qty_col}.issubset(set(df.columns)):
            raise ValueError(f"文件缺少模板指定表头：{os.path.basename(file_path)}")
        if size_col and size_col not in df.columns:
            raise ValueError(f"文件缺少模板指定尺码表头：{os.path.basename(file_path)}")

        remark_cols = find_order_remark_columns(df.columns, remark_col)
        use_cols = []
        for col in [short_col, spec_col, size_col, qty_col, raw_text_col] + remark_cols:
            if col and col not in use_cols:
                use_cols.append(col)

        temp = df[use_cols].copy()
        if template.get("name") == PROCESSED_WAYBILL_TEMPLATE_NAME:
            required_cols = [col for col in [short_col, spec_col, size_col, qty_col] if col]
            temp = temp.dropna(how="all", subset=required_cols)
        else:
            temp = temp.dropna(subset=[short_col])

        for _, row in temp.iterrows():
            names = split_items(row[short_col], item_sep)
            specs = split_items(row[spec_col], item_sep)
            sizes = split_items(row[size_col], item_sep) if size_col else []
            qtys = split_items(row[qty_col], item_sep)
            remark_text = merge_order_remark(row, remark_cols) if remark_cols else ""
            item_count = max(len(names), len(specs), len(sizes), len(qtys), 1)
            remarks = split_remark_for_items(remark_text, item_sep, item_count) if remark_text else []
            if len(remarks) > item_count and len(qtys) == 1 and normalize_qty(qtys[0]) == len(remarks):
                qtys = ["1"] * len(remarks)
            max_len = max(len(names), len(specs), len(sizes), len(qtys), len(remarks), 1)

            for i in range(max_len):
                name = names[i] if i < len(names) else (names[-1] if names else "")
                spec = specs[i] if i < len(specs) else (specs[-1] if specs else "")
                size_value = sizes[i] if i < len(sizes) else (sizes[-1] if sizes else "")
                qty = qtys[i] if i < len(qtys) else (qtys[-1] if qtys else "1")
                remark = remarks[i] if i < len(remarks) else (remarks[-1] if remarks else "")
                raw_text = row.get(raw_text_col, "") if raw_text_col else ""
                final_spec, final_size = resolve_five_field_spec_size(spec, spec_split, remark, item_sep, size_value)
                rows.append(make_five_field_item(
                    shoe=name,
                    spec=final_spec,
                    size=final_size,
                    quantity=qty,
                    remark=remark,
                    source_file=os.path.basename(file_path),
                    raw_shoe=name,
                    raw_spec=spec,
                    raw_text=raw_text,
                ).to_order_row())
    else:
        title_idx = col_letter_to_index(template.get("title_col", "S"))
        qty_idx = col_letter_to_index(template.get("qty_col", "V"))
        remark_col = template.get("remark_col", "")
        remark_idx = col_letter_to_index(remark_col) if remark_col else None
        required_idx = [title_idx, qty_idx] + ([remark_idx] if remark_idx is not None else [])
        if df.shape[1] <= max(required_idx):
            raise ValueError(f"文件列数不足：{os.path.basename(file_path)}")

        temp = df.iloc[:, required_idx].copy()
        temp.columns = ["货品标题", "数量"] + (["备注"] if remark_idx is not None else [])
        temp = temp.dropna(subset=["货品标题"])

        def guess_short_name(title):
            m = re.search(r"^(.*?)\s*颜色[:：]", str(title))
            return m.group(1).strip() if m else str(title).strip()

        for _, row in temp.iterrows():
            title = row["货品标题"]
            name = guess_short_name(title)
            qty = normalize_qty(row["数量"])
            remark_text = row.get("备注", "")
            remarks = split_remark_for_items(remark_text, item_sep, 1) if remark_text else []
            if len(remarks) > 1 and qty == len(remarks):
                qtys = ["1"] * len(remarks)
            else:
                remarks = remarks[:1] if remarks else [remark_text]
                qtys = [row["数量"]]

            for remark, item_qty in zip(remarks, qtys):
                spec = extract_raw_spec(title)
                size = extract_size(title)
                fixed_title = title
                raw_spec = spec
                remark_spec, remark_size = extract_remark_spec_size(remark, item_sep)
                if remark_spec and is_abnormal_product_spec(spec):
                    spec = remark_spec
                    fixed_title = f"{name} 颜色: {spec} 尺码: {size}"
                if remark_size and (is_weak_spec_text(size) or not is_valid_order_size(size)):
                    size = remark_size
                    fixed_title = f"{name} 颜色: {spec} 尺码: {size}"
                rows.append(make_five_field_item(
                    shoe=name,
                    spec=spec,
                    size=size,
                    quantity=item_qty,
                    remark=remark,
                    source_file=os.path.basename(file_path),
                    raw_shoe=name,
                    raw_spec=raw_spec,
                    raw_text=fixed_title,
                ).to_order_row())

    out = pd.DataFrame(rows)
    if "备注" not in out.columns:
        out["备注"] = ""
    out["数量"] = pd.to_numeric(out["数量"], errors="coerce").fillna(1).astype(int)
    return out


def extract_size(title):
    m = re.search(r"尺码[:：]\s*([0-9.]+)", str(title))
    return m.group(1) if m else "未知"


def extract_raw_spec(title):
    m = re.search(r"颜色[:：]\s*(.*?)\s*尺码", str(title))
    return normalize_text(m.group(1)) if m else "未知"


COLOR_WORDS = (
    "黑", "白", "灰", "粉", "红", "蓝", "绿", "黄", "棕", "紫",
    "橙", "米", "杏", "银", "金", "咖", "卡其", "藏青", "奶油",
)


def is_size_only_spec(spec):
    text = normalize_text(spec)
    return bool(re.fullmatch(r"(?:尺码|码数|鞋码|size)?[3-5][0-9](?:\.5)?(?:码)?", text, flags=re.I))


def is_color_only_spec(spec):
    text = normalize_text(spec)
    if not text or len(text) > 8 or any(ch.isdigit() for ch in text):
        return False
    rest = text.replace("色", "")
    for word in sorted(COLOR_WORDS, key=len, reverse=True):
        rest = rest.replace(word, "")
    return not rest


def is_weak_spec(spec):
    text = normalize_text(spec)
    return text in WEAK_SPEC_VALUES or is_size_only_spec(text) or is_abnormal_product_spec(text)


def resolve_effective_spec(row, rule_engine, matcher):
    category = row.get(SHOE_CATEGORY_COLUMN, row.get("分类", ""))
    raw_spec = normalize_text(row.get("原始规格", ""))
    cleaned_spec = rule_engine.clean_spec(raw_spec, category)
    remark_spec, _ = extract_remark_spec_size(row.get("备注", ""))
    if remark_spec and is_abnormal_product_spec(cleaned_spec):
        cleaned_spec = remark_spec
    item = matcher.find(
        category,
        "" if is_weak_spec(cleaned_spec) else cleaned_spec,
        row.get("备注", ""),
        row.get("货品标题", ""),
        row.get(RAW_SHORT_NAME_COLUMN, ""),
        row.get(SHOE_FIELD, row.get("商品简称", "")),
    )

    if item:
        image_spec = normalize_text(item.get("spec", ""))
        display_spec = normalize_image_match_text(image_spec) or image_spec
        if is_weak_spec(cleaned_spec):
            return display_spec
        if is_color_only_spec(cleaned_spec) and len(display_spec) > len(normalize_text(cleaned_spec)):
            return display_spec

    return cleaned_spec if normalize_text(cleaned_spec) else raw_spec


def get_stall(category, stall_map):
    key = normalize_text(category)
    if key in stall_map:
        return stall_map.get(key)

    key_match = normalize_match_text(key)
    if key_match:
        for candidate, stall in (stall_map or {}).items():
            if normalize_match_text(candidate) == key_match:
                return stall
        if len(key_match) >= 2:
            for candidate, stall in (stall_map or {}).items():
                candidate_match = normalize_match_text(candidate)
                if candidate_match and (key_match in candidate_match or candidate_match in key_match):
                    return stall
    return "未设置档口"


def safe_sheet_name(name):
    name = re.sub(r'[\\/:*?\[\]]', "_", str(name).strip())
    return name[:31] if name else "未设置档口"


def prepare_image_for_excel(image_b64, category, spec):
    temp_dir = get_temp_dir()
    raw = os.path.join(temp_dir, f"{safe_filename(category)}_{safe_filename(spec)}_raw")
    png = os.path.join(temp_dir, f"{safe_filename(category)}_{safe_filename(spec)}.png")
    base64_to_image_file(image_b64, raw)
    return prepare_image_file_for_excel(raw, category, spec)


def prepare_image_file_for_excel(source_path, category, spec):
    temp_dir = get_temp_dir()
    png = os.path.join(temp_dir, f"{safe_filename(category)}_{safe_filename(spec)}.png")

    img = PILImage.open(source_path)
    if img.mode in ("RGBA", "LA"):
        bg = PILImage.new("RGBA", img.size, (255, 255, 255, 255))
        bg.paste(img, mask=img.getchannel("A"))
        img = bg.convert("RGB")
    else:
        img = img.convert("RGB")

    img.thumbnail((IMAGE_WIDTH_PX, IMAGE_HEIGHT_PX), PILImage.LANCZOS)
    canvas = PILImage.new("RGB", (IMAGE_WIDTH_PX, IMAGE_HEIGHT_PX), (255, 255, 255))
    canvas.paste(img, ((IMAGE_WIDTH_PX - img.width) // 2, (IMAGE_HEIGHT_PX - img.height) // 2))
    canvas.save(png)
    return png


def prepare_image_item_for_excel(item, category, spec, image_cache=None):
    if not item:
        return ""

    cache_key = make_image_key(category, spec)
    if image_cache is not None and cache_key in image_cache:
        return image_cache[cache_key]

    prepared = ""
    image_file = item.get("image_file") or item.get("file") or ""
    if image_file:
        source = image_file
        if not os.path.isabs(source):
            source = os.path.join(get_data_dir(), image_file)
        if os.path.exists(source):
            prepared = prepare_image_file_for_excel(source, category, spec)

    image_b64 = item.get("image_base64")
    if not prepared and image_b64:
        prepared = prepare_image_for_excel(image_b64, category, spec)

    if image_cache is not None:
        image_cache[cache_key] = prepared

    return prepared


def style_and_images(output_file, image_map, image_cache=None):
    matcher = ImageMatcher(image_map)
    wb = load_workbook(output_file)
    for ws in wb.worksheets:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = {normalize_text(ws.cell(row=1, column=col).value): col for col in range(1, ws.max_column + 1)}
        category_col = headers.get(SHOE_CATEGORY_COLUMN) or headers.get("分类", 1)
        spec_col = headers.get("规格", 2)
        image_col = headers.get("图片", 3)
        remark_col = headers.get("备注")
        match_col = headers.get(MATCH_TEXT_COLUMN)
        image_col_letter = ws.cell(row=1, column=image_col).column_letter
        if match_col:
            ws.column_dimensions[ws.cell(row=1, column=match_col).column_letter].hidden = True

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for col in range(1, ws.max_column + 1):
            col_letter = ws.cell(row=1, column=col).column_letter
            header = normalize_text(ws.cell(row=1, column=col).value)
            if header == "图片":
                ws.column_dimensions[col_letter].width = 20
            elif header == "尺码":
                ws.column_dimensions[col_letter].width = 42
            elif header == "规格":
                ws.column_dimensions[col_letter].width = 28
            else:
                ws.column_dimensions[col_letter].width = 16

        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = IMAGE_HEIGHT_PX * 0.75
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            category = normalize_text(ws.cell(row=row, column=category_col).value)
            spec = normalize_text(ws.cell(row=row, column=spec_col).value)
            remark = ws.cell(row=row, column=remark_col).value if remark_col else ""
            match_text = ws.cell(row=row, column=match_col).value if match_col else ""
            item = matcher.find(category, spec, remark, match_text)

            if item:
                try:
                    png_path = prepare_image_item_for_excel(item, category, spec, image_cache)
                    if png_path:
                        img = XLImage(png_path)
                        img.width = IMAGE_WIDTH_PX
                        img.height = IMAGE_HEIGHT_PX
                        ws.add_image(img, f"{image_col_letter}{row}")
                        ws.cell(row=row, column=image_col).value = ""
                except Exception:
                    ws.cell(row=row, column=image_col).value = ""
    wb.save(output_file)


def is_waybill_order_template(template):
    name = str((template or {}).get("name", "") or "").strip()
    mode = str((template or {}).get("mode", "") or "").strip()
    return name in {WAYBILL_RAW_TEMPLATE_NAME, PROCESSED_WAYBILL_TEMPLATE_NAME} or mode == RAW_WAYBILL_MODE


def build_result(files, system, template_name=""):
    template = get_template(system, template_name)
    waybill_input = is_waybill_order_template(template)
    rules = list(system.get("category_rules", []) or [])
    if not rules and not waybill_input:
        raise ValueError("当前整理系统未配置鞋款分类规则，请先在管理系统里配置")

    rule_engine = RuleEngine(rules)

    df = pd.concat([read_by_template(f, template, system.get("waybill_parse_rules", {})) for f in files], ignore_index=True)
    if SPEC_FIELD in df.columns:
        df["原始规格"] = df[SPEC_FIELD].apply(normalize_text)
    elif RAW_SPEC_FIELD in df.columns:
        df["原始规格"] = df[RAW_SPEC_FIELD].apply(normalize_text)
    else:
        df["原始规格"] = df["货品标题"].apply(extract_raw_spec)
    if "备注" not in df.columns:
        df["备注"] = ""
    if SIZE_FIELD not in df.columns:
        df[SIZE_FIELD] = df["货品标题"].apply(extract_size)
    df["尺码"] = df.apply(lambda r: resolve_effective_size(r.get(SIZE_FIELD, ""), r.get("备注", ""), r.get("原始规格", "")), axis=1)
    if SHOE_FIELD not in df.columns:
        df[SHOE_FIELD] = df["商品简称"]
    df[RAW_SHORT_NAME_COLUMN] = df.get(RAW_SHOE_FIELD, df[SHOE_FIELD])

    if waybill_input:
        df[SHOE_CATEGORY_COLUMN] = df[SHOE_FIELD].apply(lambda v: clean_short_name_for_output(v) or "未分类")
    else:
        df[SHOE_CATEGORY_COLUMN] = df.apply(
            lambda r: rule_engine.detect_category(
                r.get(SHOE_FIELD, r.get("商品简称", "")),
                r["货品标题"],
                r["原始规格"],
                r.get("备注", ""),
                r.get("尺码", ""),
                r.get("数量", ""),
            ),
            axis=1,
        )
    # 兼容旧图片/档口代码中仍使用“分类”的内部命名。
    df["分类"] = df[SHOE_CATEGORY_COLUMN]
    if waybill_input:
        df[SHOE_FIELD] = df[SHOE_CATEGORY_COLUMN]
    else:
        df[SHOE_FIELD] = df.apply(
            lambda r: rule_engine.detect_output_shoe(
                r.get(RAW_SHORT_NAME_COLUMN, ""),
                r.get("货品标题", ""),
                r.get("原始规格", ""),
                r.get("备注", ""),
                r.get("尺码", ""),
                r.get("数量", ""),
                r.get(SHOE_CATEGORY_COLUMN, ""),
            )
            or clean_short_name_for_output(r.get(RAW_SHORT_NAME_COLUMN, ""), r.get(SHOE_CATEGORY_COLUMN, "")),
            axis=1,
        )
    df["商品简称"] = df[SHOE_FIELD]
    used_categories = sorted({normalize_text(c) for c in df[SHOE_CATEGORY_COLUMN].dropna().tolist() if normalize_text(c)})
    image_map = load_image_map_for_categories(system, used_categories)
    matcher = ImageMatcher(image_map)
    df["规格"] = df.apply(lambda r: resolve_effective_spec(r, rule_engine, matcher), axis=1)
    df["档口"] = df[SHOE_CATEGORY_COLUMN].apply(lambda c: get_stall(c, system.get("stall_map", {})))
    df[MATCH_TEXT_COLUMN] = df.apply(
        lambda r: merge_text_values([
            r.get("规格", ""),
            r.get("原始规格", ""),
            r.get("备注", ""),
            strip_shop_codes(r.get("货品标题", "")),
            r.get("货品标题", ""),
            r.get(SHOE_FIELD, ""),
        ]),
        axis=1
    )

    result = (
        df.groupby(["档口", SHOE_CATEGORY_COLUMN, "规格"], dropna=False)
        .apply(lambda g: pd.Series({
            "尺码": merge_size_quantity(g),
            "数量": sum(normalize_qty(v) for v in g["数量"]),
            "备注": merge_text_values(g["备注"]),
            MATCH_TEXT_COLUMN: merge_text_values(g[MATCH_TEXT_COLUMN], max_chars=600),
        }))
        .reset_index()
        .sort_values(by=["档口", SHOE_CATEGORY_COLUMN, "规格"])
    )
    result.attrs["image_map"] = image_map
    return result


def generate_order_file(files, system, output_mode="按档口分Sheet", template_name=""):
    result = build_result(files, system, template_name)
    image_map = result.attrs.get("image_map", {})
    image_cache = {}
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = get_output_dir()

    if output_mode == "合并一个Sheet":
        out_file = os.path.join(out_dir, f"订单整理文档_合并_{ts}.xlsx")
        out_df = result[["档口", SHOE_CATEGORY_COLUMN, "规格", "尺码", "数量", "备注", MATCH_TEXT_COLUMN]].copy()
        out_df.insert(3, "图片", "")
        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            out_df.to_excel(writer, sheet_name="全部订单", index=False)
        style_and_images(out_file, image_map, image_cache)
        return out_file

    if output_mode == "按档口分文档":
        batch = os.path.join(out_dir, f"订单整理文档_按档口分文档_{ts}")
        os.makedirs(batch, exist_ok=True)
        for stall, stall_df in result.groupby("档口", dropna=False):
            out_file = os.path.join(batch, f"{safe_sheet_name(stall)}_{ts}.xlsx")
            out_df = stall_df[[SHOE_CATEGORY_COLUMN, "规格", "尺码", "数量", "备注", MATCH_TEXT_COLUMN]].copy()
            out_df.insert(2, "图片", "")
            with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
                out_df.to_excel(writer, sheet_name=safe_sheet_name(stall), index=False)
            style_and_images(out_file, image_map, image_cache)
        return batch

    out_file = os.path.join(out_dir, f"订单整理文档_分Sheet_{ts}.xlsx")
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        for stall, stall_df in result.groupby("档口", dropna=False):
            out_df = stall_df[[SHOE_CATEGORY_COLUMN, "规格", "尺码", "数量", "备注", MATCH_TEXT_COLUMN]].copy()
            out_df.insert(2, "图片", "")
            out_df.to_excel(writer, sheet_name=safe_sheet_name(stall), index=False)
    style_and_images(out_file, image_map, image_cache)
    return out_file
