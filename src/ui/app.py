import os
import json
import sys
import time
import traceback
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from typing import List, Optional

SRC_ROOT = Path(__file__).resolve().parents[1]
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from fastapi import FastAPI, File, Form, Header, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse

from core import collector_agent_store
from core.order_core import generate_order_file
from core.waybill_raw_pipeline import parse_raw_waybill_records, write_processed_waybill_xlsx
from utils.app_info import APP_VERSION
from utils.order_secure_common import (
    WAYBILL_PROCESSED_TEMPLATE_NAME,
    WAYBILL_TEMPLATE_NAME,
    get_data_dir,
    get_data_file,
    get_output_dir,
    image_storage_summary,
    load_data,
    load_templates_fast,
)
from core import waybill_files


app = FastAPI(title="订单整理系统 Web服务")

WEB_VERSION = APP_VERSION
ALLOWED_OUTPUT_MODES = {"合并一个Sheet", "按档口分Sheet", "按档口分文档"}
WAYBILL_REMOTE_STATE = {
    "status": "idle",
    "batch_id": "",
    "started_at": "",
    "stopped_at": "",
    "last_raw_file": "",
    "last_raw_count": 0,
    "last_processed_file": "",
    "last_processed_count": 0,
    "collectors": {},
    "uploads": {},
}
WAYBILL_COLLECTOR_ONLINE_SECONDS = 20
WAYBILL_STOP_WAIT_SECONDS = 8

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(get_output_dir(), "_web_uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
DATA_CACHE = {"signature": None, "data": None}


def debug_environment():
    info = {
        "web_version": WEB_VERSION,
        "app_file": __file__,
        "base_dir": BASE_DIR,
        "data_dir": get_data_dir(),
        "data_file": get_data_file(),
    }
    try:
        from core import order_core

        info["order_core_file"] = getattr(order_core, "__file__", "")
        info["order_core_has_generate_order_file"] = hasattr(order_core, "generate_order_file")
        info["order_core_has_five_field_module"] = hasattr(order_core, "SHOE_FIELD")
    except Exception:
        info["order_core_import_error"] = traceback.format_exc()

    try:
        from core import five_field_normalizer

        info["five_field_normalizer_file"] = getattr(five_field_normalizer, "__file__", "")
        info["five_fields"] = getattr(five_field_normalizer, "FIVE_FIELDS", [])
    except Exception:
        info["five_field_normalizer_error"] = traceback.format_exc()

    try:
        from utils import order_secure_common

        info["order_secure_common_file"] = getattr(order_secure_common, "__file__", "")
        info["secure_common_has_get_data_dir"] = hasattr(order_secure_common, "get_data_dir")
    except Exception:
        info["order_secure_common_import_error"] = traceback.format_exc()
    return info


def load_html(name):
    frozen_root = getattr(sys, "_MEIPASS", "")
    meipass = Path(frozen_root) if frozen_root else None
    candidates = [
        meipass / "ui" / "templates" / name if meipass else None,
        meipass / "templates" / name if meipass else None,
        Path(BASE_DIR) / "templates" / name,
    ]
    for path in candidates:
        if path and path.exists():
            with path.open("r", encoding="utf-8") as f:
                return f.read()
    with open(os.path.join(BASE_DIR, "templates", name), "r", encoding="utf-8") as f:
        return f.read()


def get_data_signature():
    path = get_data_file()
    if not os.path.exists(path):
        return None
    stat = os.stat(path)
    return (stat.st_mtime_ns, stat.st_size)


def load_runtime_data():
    signature = get_data_signature()
    if DATA_CACHE["data"] is not None and DATA_CACHE["signature"] == signature:
        return DATA_CACHE["data"]

    data = load_data(auto_save_on_read=False)
    DATA_CACHE["signature"] = get_data_signature()
    DATA_CACHE["data"] = data
    return data


def current_time_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def parse_time_text(value):
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt)
        except Exception:
            continue
    return None


def collector_is_online(collector, max_age_seconds=WAYBILL_COLLECTOR_ONLINE_SECONDS):
    seen = parse_time_text(collector.get("last_seen"))
    if not seen:
        return False
    return (datetime.now() - seen).total_seconds() <= max_age_seconds


def public_waybill_collectors():
    uploads = WAYBILL_REMOTE_STATE.get("uploads", {})
    collector_items = WAYBILL_REMOTE_STATE.get("collectors", {})
    persisted_agents = {row.get("client_id"): row for row in collector_agent_store.list_agents_public() if row.get("client_id")}
    client_ids = sorted(set(persisted_agents.keys()) | set(collector_items.keys()) | set(uploads.keys()))
    rows = []
    for client_id in client_ids:
        agent = persisted_agents.get(client_id, {})
        item = collector_items.get(client_id, {})
        upload = uploads.get(client_id, {})
        components = (
            agent.get("component_status")
            if isinstance(agent.get("component_status"), list)
            else item.get("component_status")
            if isinstance(item.get("component_status"), list)
            else item.get("components")
            if isinstance(item.get("components"), list)
            else []
        )
        rows.append(
            {
                "client_id": client_id,
                "machine_name": agent.get("machine_name") or item.get("machine_name") or upload.get("machine_name") or "",
                "machine_label": agent.get("machine_label") or item.get("machine_label") or upload.get("machine_label") or client_id,
                "hostname": agent.get("hostname") or item.get("hostname") or upload.get("hostname") or "",
                "username": agent.get("username") or item.get("username") or upload.get("username") or "",
                "platform": agent.get("platform") or item.get("platform") or "",
                "last_seen": agent.get("last_seen") or item.get("last_seen") or "",
                "online": bool(agent.get("online")) or collector_is_online(item),
                "components": components,
                "component_status": components,
                "component_count": len(components),
                "available_components": len([row for row in components if row.get("exists")]),
                "uploaded": bool(upload),
                "uploaded_records": len(upload.get("records", [])) if upload else int(agent.get("last_upload_count") or 0),
                "uploaded_at": upload.get("uploaded_at") or agent.get("last_upload_at") or "",
                "records_found": upload.get("records_found", 0) if upload else 0,
                "agent_version": agent.get("agent_version") or item.get("agent_version") or upload.get("agent_version") or "",
                "protocol_version": agent.get("protocol_version") or item.get("protocol_version") or upload.get("protocol_version") or "",
                "upgrade_required": bool(agent.get("upgrade_required")),
                "upgrade_message": agent.get("upgrade_message") or "",
                "download_url": agent.get("download_url") or "",
            }
        )
    return rows


def online_waybill_collector_ids():
    return [
        row.get("client_id")
        for row in public_waybill_collectors()
        if row.get("client_id") and row.get("online")
    ]


def dedupe_waybill_records(records):
    merged = []
    keyed_positions = {}
    for row in records:
        key = waybill_files.record_key(row)
        if key:
            previous = keyed_positions.get(key)
            if previous is not None:
                merged[previous] = row
                continue
            keyed_positions[key] = len(merged)
        merged.append(row)
    return merged


def uploaded_waybill_records():
    records = []
    uploads = WAYBILL_REMOTE_STATE.get("uploads", {})
    for client_id, upload in uploads.items():
        for row in upload.get("records", []):
            record = dict(row)
            record.setdefault("source_client_id", client_id)
            record.setdefault("machine_name", upload.get("machine_name", ""))
            record.setdefault("machine_label", upload.get("machine_label", ""))
            records.append(record)
    return records


def finalize_remote_waybill_batch():
    records = uploaded_waybill_records()
    batch_tag = waybill_files.safe_batch_tag(WAYBILL_REMOTE_STATE.get("batch_id") or WAYBILL_REMOTE_STATE.get("started_at"))
    if records:
        export_result = waybill_files.export_records(records, merge_existing=False, batch_tag=batch_tag)
        raw_path = waybill_files.unique_path(waybill_files.raw_waybill_path(batch_tag=batch_tag))
        raw_file = str(waybill_files.write_raw_waybill_xlsx(records, raw_path))
        data = load_data(auto_save_on_read=False)
        system = data.get("systems", {}).get(data.get("active_system", "default"), {}) if isinstance(data, dict) else {}
        processed_rows = parse_raw_waybill_records(records, system.get("waybill_parse_rules", {}))
        processed_path = waybill_files.unique_path(waybill_files.processed_waybill_path(batch_tag=batch_tag))
        processed_file = str(write_processed_waybill_xlsx(processed_rows, processed_path))
    else:
        export_result = {"total": 0, "added": 0, "xlsx": "", "jsonl": ""}
        raw_file = ""
        processed_rows = []
        processed_file = ""

    WAYBILL_REMOTE_STATE["last_raw_file"] = raw_file
    WAYBILL_REMOTE_STATE["last_raw_count"] = len(records)
    WAYBILL_REMOTE_STATE["last_processed_file"] = processed_file
    WAYBILL_REMOTE_STATE["last_processed_count"] = len(processed_rows)
    return records, raw_file, processed_file, export_result


def wait_for_waybill_uploads(expected_client_ids):
    deadline = time.time() + WAYBILL_STOP_WAIT_SECONDS
    expected = {item for item in expected_client_ids if item}
    while time.time() < deadline:
        uploads = WAYBILL_REMOTE_STATE.get("uploads", {})
        if not expected or expected.issubset(set(uploads.keys())):
            break
        time.sleep(0.4)


def latest_waybill_raw_file():
    state_path = str(WAYBILL_REMOTE_STATE.get("last_raw_file") or "").strip()
    if state_path and os.path.exists(state_path) and not os.path.isdir(state_path):
        return state_path
    if str(WAYBILL_REMOTE_STATE.get("status") or "") in {"running", "stopping"}:
        return ""


def latest_waybill_processed_file():
    state_path = str(WAYBILL_REMOTE_STATE.get("last_processed_file") or "").strip()
    if state_path and os.path.exists(state_path) and not os.path.isdir(state_path):
        return state_path
    if str(WAYBILL_REMOTE_STATE.get("status") or "") in {"running", "stopping"}:
        return ""

    try:
        output_dir = waybill_files.get_waybill_output_dir()
        candidates = sorted(output_dir.glob("监控面单识别_*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
        return str(candidates[0]) if candidates else ""
    except Exception:
        return ""

    try:
        output_dir = waybill_files.get_waybill_output_dir()
        candidates = sorted(output_dir.glob("监控面单原文_*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
        return str(candidates[0]) if candidates else ""
    except Exception:
        return ""


def waybill_raw_row_count(path):
    if not path or not os.path.exists(path):
        return 0
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        return max(0, (ws.max_row or 1) - 1)
    except Exception:
        return 0


def waybill_status_payload():
    collectors = public_waybill_collectors()
    online_collectors = [item for item in collectors if item.get("online")]
    records = uploaded_waybill_records()
    status_text = str(WAYBILL_REMOTE_STATE.get("status") or "idle")
    last_raw_file = latest_waybill_raw_file()
    last_raw_count = WAYBILL_REMOTE_STATE.get("last_raw_count", 0) or waybill_raw_row_count(last_raw_file)
    last_processed_file = latest_waybill_processed_file()
    last_processed_count = WAYBILL_REMOTE_STATE.get("last_processed_count", 0) or waybill_raw_row_count(last_processed_file)
    return {
        "ok": True,
        "web_version": WEB_VERSION,
        "active": status_text in {"running", "stopping"},
        "status": status_text,
        "batch_id": WAYBILL_REMOTE_STATE.get("batch_id", ""),
        "started_at": WAYBILL_REMOTE_STATE.get("started_at", ""),
        "stopped_at": WAYBILL_REMOTE_STATE.get("stopped_at", ""),
        "last_raw_file": last_raw_file,
        "last_raw_filename": os.path.basename(last_raw_file) if last_raw_file else "",
        "last_raw_count": last_raw_count,
        "last_processed_file": last_processed_file,
        "last_processed_filename": os.path.basename(last_processed_file) if last_processed_file else "",
        "last_processed_count": last_processed_count,
        "records": len(records),
        "session_records": len(records) if status_text in {"running", "stopping"} else last_raw_count,
        "collectors": collectors,
        "online_collectors": len(online_collectors),
        "collector_count": len(collectors),
        "uploaded_collectors": len(WAYBILL_REMOTE_STATE.get("uploads", {})),
        "template_name": WAYBILL_TEMPLATE_NAME,
        "processed_template_name": WAYBILL_PROCESSED_TEMPLATE_NAME,
        "collector_version": collector_agent_store.LATEST_AGENT_VERSION,
        "collector_protocol_version": collector_agent_store.COLLECTOR_PROTOCOL_VERSION,
        "raw_records": collector_agent_store.list_raw_records(limit=10),
    }


def parse_server_files(raw_value):
    if not raw_value:
        return []
    try:
        values = json.loads(raw_value)
    except Exception:
        values = [raw_value]
    if isinstance(values, str):
        values = [values]
    if not isinstance(values, list):
        raise ValueError("server_files 格式无效")

    result = []
    for value in values:
        path = os.path.abspath(str(value or "").strip())
        if not path:
            continue
        if not is_path_under(path, get_output_dir()):
            raise ValueError("只能使用输出目录内的服务器文件")
        if not os.path.exists(path) or os.path.isdir(path):
            raise ValueError(f"服务器文件不存在：{os.path.basename(path)}")
        result.append(path)
    return result


def is_path_under(path, root):
    try:
        path_abs = os.path.abspath(path)
        root_abs = os.path.abspath(root)
        return os.path.commonpath([path_abs, root_abs]) == root_abs
    except Exception:
        return False


def zip_output_folder(folder_path):
    folder_path = os.path.abspath(folder_path)
    if not is_path_under(folder_path, get_output_dir()) or not os.path.isdir(folder_path):
        raise ValueError("只能打包输出目录内的文件夹")

    zip_path = folder_path.rstrip("\\/") + ".zip"
    if os.path.exists(zip_path):
        os.remove(zip_path)

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(folder_path):
            for filename in files:
                file_path = os.path.join(root, filename)
                arcname = os.path.relpath(file_path, os.path.dirname(folder_path))
                zf.write(file_path, arcname)

    return zip_path


def get_current_system():
    data = load_runtime_data()

    if isinstance(data, dict) and "systems" in data:
        system_id = data.get("active_system", "default")
        systems = data.get("systems", {})

        if system_id not in systems and systems:
            system_id = next(iter(systems.keys()))

        system = systems.get(system_id, {})
        return system, system_id

    return data, "default"


@app.post("/api/collector/register")
def api_collector_register(payload: dict):
    result, error = collector_agent_store.register_agent(payload)
    if error:
        return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": error}, status_code=400)
    return {"ok": True, "web_version": WEB_VERSION, **(result or {}), **collector_agent_store.version_info()}


@app.post("/api/collector/poll")
def api_collector_poll(payload: dict, authorization: str = Header(default="")):
    agent, error = collector_agent_store.authenticate_agent(payload, authorization)
    if error:
        return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": error}, status_code=401)

    agent = collector_agent_store.upsert_agent_from_poll(payload, agent or {})
    client_id = str(agent.get("client_id") or payload.get("client_id") or "")
    collectors = WAYBILL_REMOTE_STATE.setdefault("collectors", {})
    collectors[client_id] = {
        "client_id": client_id,
        "machine_name": str(payload.get("machine_name") or agent.get("machine_name") or ""),
        "machine_label": str(payload.get("machine_label") or agent.get("machine_label") or client_id),
        "hostname": str(payload.get("hostname") or agent.get("hostname") or ""),
        "username": str(payload.get("username") or agent.get("username") or ""),
        "platform": str(payload.get("platform") or agent.get("platform") or ""),
        "active_batch_id": str(payload.get("active_batch_id") or ""),
        "component_status": payload.get("component_status") if isinstance(payload.get("component_status"), list) else [],
        "agent_version": str(payload.get("agent_version") or agent.get("agent_version") or ""),
        "protocol_version": str(payload.get("protocol_version") or agent.get("protocol_version") or ""),
        "last_seen": current_time_text(),
    }

    status_text = str(WAYBILL_REMOTE_STATE.get("status") or "idle")
    batch_id = str(WAYBILL_REMOTE_STATE.get("batch_id") or "")
    uploads = WAYBILL_REMOTE_STATE.get("uploads", {})
    if collector_agent_store.agent_needs_upgrade(agent.get("agent_version"), agent.get("protocol_version")):
        command = "upgrade"
    elif status_text == "running" and batch_id:
        command = "start"
    elif status_text in {"stopping", "finished"} and batch_id and client_id not in uploads:
        command = "stop"
    else:
        command = "idle"

    version = collector_agent_store.version_info()
    version["upgrade_required"] = command == "upgrade"
    if command == "upgrade":
        version["upgrade_message"] = "业务机采集助手版本或协议不兼容，请升级后继续采集。"
    return {
        "ok": True,
        "web_version": WEB_VERSION,
        "command": command,
        "batch_id": batch_id,
        "status": status_text,
        "poll_interval_seconds": 2,
        **version,
    }


@app.post("/api/collector/upload")
def api_collector_upload(payload: dict, authorization: str = Header(default="")):
    agent, error = collector_agent_store.authenticate_agent(payload, authorization)
    if error:
        return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": error}, status_code=401)

    client_id = str(payload.get("client_id") or "").strip()
    batch_id = str(payload.get("batch_id") or "").strip()
    if not client_id or not batch_id:
        return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": "client_id_and_batch_id_required"}, status_code=400)
    if batch_id != str(WAYBILL_REMOTE_STATE.get("batch_id") or ""):
        return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": "batch_id_not_current"}, status_code=409)

    records_raw = payload.get("records", [])
    if not isinstance(records_raw, list):
        return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": "records_must_be_list"}, status_code=400)
    try:
        accepted_records = collector_agent_store.append_raw_records(records_raw, payload, agent or {})
    except ValueError as exc:
        return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": str(exc)}, status_code=400)

    machine_name = str(payload.get("machine_name") or agent.get("machine_name") or "")
    machine_label = str(payload.get("machine_label") or agent.get("machine_label") or machine_name or client_id)
    WAYBILL_REMOTE_STATE.setdefault("uploads", {})[client_id] = {
        "client_id": client_id,
        "machine_name": machine_name,
        "machine_label": machine_label,
        "hostname": str(payload.get("hostname") or agent.get("hostname") or ""),
        "username": str(payload.get("username") or agent.get("username") or ""),
        "uploaded_at": current_time_text(),
        "records_found": int(payload.get("records_found") or len(records_raw)),
        "agent_version": str(payload.get("agent_version") or agent.get("agent_version") or ""),
        "protocol_version": str(payload.get("protocol_version") or agent.get("protocol_version") or ""),
        "records": accepted_records,
    }
    if WAYBILL_REMOTE_STATE.get("status") in {"stopping", "finished"}:
        finalize_remote_waybill_batch()
    return {
        "ok": True,
        "web_version": WEB_VERSION,
        "batch_id": batch_id,
        "received": len(records_raw),
        "accepted": len(accepted_records),
        "uploaded_collectors": list(WAYBILL_REMOTE_STATE.get("uploads", {}).keys()),
    }


@app.get("/api/collector/agents")
def api_collector_agents():
    return {"ok": True, "web_version": WEB_VERSION, "agents": public_waybill_collectors()}


@app.get("/api/collector/records")
def api_collector_records(limit: int = 50, record_id: str = ""):
    if record_id:
        record = collector_agent_store.get_raw_record(record_id)
        if not record:
            return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": "record_not_found"}, status_code=404)
        return {"ok": True, "web_version": WEB_VERSION, "record": record}
    return {"ok": True, "web_version": WEB_VERSION, "records": collector_agent_store.list_raw_records(limit=limit)}


@app.get("/api/collector/records/{record_id}")
def api_collector_record_detail(record_id: str):
    record = collector_agent_store.get_raw_record(record_id)
    if not record:
        return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": "record_not_found"}, status_code=404)
    return {"ok": True, "web_version": WEB_VERSION, "record": record}


@app.get("/api/collector/version-info")
def api_collector_version_info():
    return {"ok": True, "web_version": WEB_VERSION, **collector_agent_store.version_info()}


@app.get("/", response_class=HTMLResponse)
def index():
    return load_html("index.html")


@app.get("/api/version")
def api_version():
    return {"ok": True, "web_version": WEB_VERSION}


@app.get("/api/self-check")
def api_self_check():
    info = debug_environment()
    info["ok"] = (
        info.get("order_core_has_generate_order_file") is True
        and info.get("order_core_has_five_field_module") is True
        and bool(info.get("five_fields"))
    )
    return info


@app.get("/api/debug/core-check")
def api_debug_core_check():
    try:
        from core import five_field_normalizer
        from core.order_core import generate_order_file as _generate_order_file

        return {
            "ok": True,
            "web_version": WEB_VERSION,
            "message": "五要素模块和订单核心模块导入正常",
            "five_fields": getattr(five_field_normalizer, "FIVE_FIELDS", []),
            "debug": debug_environment(),
        }
    except Exception:
        return JSONResponse(
            {
                "ok": False,
                "web_version": WEB_VERSION,
                "error": "核心模块导入失败",
                "traceback": traceback.format_exc(),
                "debug": debug_environment(),
            },
            status_code=500,
        )


@app.get("/api/status")
def status():
    system, system_id = get_current_system()
    image_stats = image_storage_summary(count_entries=True)

    return {
        "ok": True,
        "web_version": WEB_VERSION,
        "system_id": system_id,
        "system_name": system.get("name", system_id),
        "active_template": system.get("active_template", ""),
        "category_rules": len(system.get("category_rules", [])),
        "stall_rules": len(system.get("stall_map", {})),
        "image_rules": image_stats.get("entries", 0) or 0,
        "image_category_files": image_stats.get("category_files", 0),
        "image_storage_mb": round((image_stats.get("bytes", 0) or 0) / 1024 / 1024, 2),
    }


@app.get("/api/templates")
def api_templates():
    system, _ = get_current_system()
    templates = load_templates_fast() or system.get("import_templates", [])
    active_template = system.get("active_template", "")

    return {
        "ok": True,
        "web_version": WEB_VERSION,
        "active_template": active_template,
        "templates": [
            {
                "name": t.get("name", ""),
                "mode": t.get("mode", ""),
            }
            for t in templates
            if t.get("name", "")
        ],
    }


@app.get("/api/waybill/status")
def api_waybill_status():
    return waybill_status_payload()


@app.post("/api/waybill/start")
def api_waybill_start():
    batch_id = f"WB{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"
    WAYBILL_REMOTE_STATE["status"] = "running"
    WAYBILL_REMOTE_STATE["batch_id"] = batch_id
    WAYBILL_REMOTE_STATE["started_at"] = current_time_text()
    WAYBILL_REMOTE_STATE["stopped_at"] = ""
    WAYBILL_REMOTE_STATE["last_raw_file"] = ""
    WAYBILL_REMOTE_STATE["last_raw_count"] = 0
    WAYBILL_REMOTE_STATE["last_processed_file"] = ""
    WAYBILL_REMOTE_STATE["last_processed_count"] = 0
    WAYBILL_REMOTE_STATE["uploads"] = {}
    return waybill_status_payload()


@app.post("/api/waybill/stop")
def api_waybill_stop():
    if WAYBILL_REMOTE_STATE.get("status") not in {"running", "stopping"}:
        return JSONResponse({"ok": False, "error": "打印机监听未启动"}, status_code=400)

    expected_client_ids = online_waybill_collector_ids()
    WAYBILL_REMOTE_STATE["status"] = "stopping"
    WAYBILL_REMOTE_STATE["stopped_at"] = current_time_text()
    wait_for_waybill_uploads(expected_client_ids)
    records, raw_file, processed_file, export_result = finalize_remote_waybill_batch()
    WAYBILL_REMOTE_STATE["status"] = "finished"

    payload = waybill_status_payload()
    payload.update(
        {
            "records_found": len(records),
            "session_records": len(records),
            "raw_file": raw_file,
            "raw_filename": os.path.basename(raw_file) if raw_file else "",
            "processed_file": processed_file,
            "processed_filename": os.path.basename(processed_file) if processed_file else "",
            "processed_records": WAYBILL_REMOTE_STATE.get("last_processed_count", 0),
            "processed_template_name": WAYBILL_PROCESSED_TEMPLATE_NAME,
            "export_result": export_result,
            "expected_collectors": expected_client_ids,
            "uploaded_collectors": list(WAYBILL_REMOTE_STATE.get("uploads", {}).keys()),
        }
    )
    return payload


@app.post("/api/waybill/agent/poll")
def api_waybill_agent_poll(payload: dict, authorization: str = Header(default="")):
    return api_collector_poll(payload, authorization)


@app.post("/api/waybill/agent/upload")
def api_waybill_agent_upload(payload: dict, authorization: str = Header(default="")):
    return api_collector_upload(payload, authorization)


@app.post("/api/generate")
async def api_generate(
    files: Optional[List[UploadFile]] = File(None),
    output_mode: str = Form("按档口分Sheet"),
    template_name: str = Form(""),
    server_files: str = Form(""),
):
    system, _ = get_current_system()
    system = dict(system or {})
    system["import_templates"] = load_templates_fast() or system.get("import_templates", [])

    if not system:
        return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": "未配置整理系统", "debug": debug_environment()}, status_code=403)

    if not template_name:
        return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": "请先选择导入模板", "debug": debug_environment()}, status_code=400)

    if output_mode not in ALLOWED_OUTPUT_MODES:
        return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": "输出方式无效", "debug": debug_environment()}, status_code=400)

    saved = []
    server_saved = []
    try:
        server_saved = parse_server_files(server_files)
        for file in files or []:
            if not file or not file.filename:
                continue
            suffix = os.path.splitext(file.filename)[1] or ".xlsx"
            path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4().hex}{suffix}")
            with open(path, "wb") as f:
                f.write(await file.read())
            saved.append(path)

        generate_files = saved + server_saved
        if not generate_files:
            return JSONResponse({"ok": False, "web_version": WEB_VERSION, "error": "请先选择订单 Excel，或先用监听打印机导入本次面单", "debug": debug_environment()}, status_code=400)

        output = generate_order_file(generate_files, system, output_mode, template_name)
        filename = os.path.basename(output)
        is_dir = os.path.isdir(output)
        download_path = zip_output_folder(output) if is_dir else output
        return {
            "ok": True,
            "web_version": WEB_VERSION,
            "path": output,
            "filename": filename,
            "is_dir": is_dir,
            "download_path": download_path,
            "download_filename": os.path.basename(download_path),
            "debug": debug_environment(),
        }
    except Exception as exc:
        return JSONResponse(
            {
                "ok": False,
                "web_version": WEB_VERSION,
                "error": str(exc),
                "traceback": traceback.format_exc(),
                "debug": debug_environment(),
            },
            status_code=400,
        )
    finally:
        for path in saved:
            try:
                if os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass


@app.get("/api/download")
def download(path: str):
    if not is_path_under(path, get_output_dir()):
        return JSONResponse({"ok": False, "error": "只能下载输出目录内的文件"}, status_code=403)
    if not os.path.exists(path) or os.path.isdir(path):
        return JSONResponse({"ok": False, "error": "文件不存在或不是单个文件"}, status_code=404)
    return FileResponse(path, filename=os.path.basename(path))


@app.get("/api/open-output")
def open_output():
    os.startfile(get_output_dir())
    return {"ok": True}
