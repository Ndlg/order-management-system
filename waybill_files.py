# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import os
import re
import socket
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from waybill_raw_contract import RAW_WAYBILL_TEXT_COLUMN, RAW_WAYBILL_TRACKING_FIELDS
from order_secure_common import get_output_dir


FIELDS = [
    ("task_id", "任务ID"),
    ("document_id", "文档ID"),
    ("task_time", "任务时间"),
    ("source_client_id", "采集端ID"),
    ("source_machine", "来源机器"),
    ("source_record_index", "来源序号"),
    ("print_text", "打印信息"),
]

RAW_WAYBILL_HEADERS = RAW_WAYBILL_TRACKING_FIELDS + [RAW_WAYBILL_TEXT_COLUMN]


def safe_filename(value: str) -> str:
    value = re.sub(r'[<>:"/\\|?*\x00-\x1F]+', "_", str(value)).strip(" ._")
    return value or "本机"


MACHINE_NAME = safe_filename(socket.gethostname())


def get_waybill_output_dir() -> Path:
    path = Path(get_output_dir()) / "waybill-monitor"
    path.mkdir(parents=True, exist_ok=True)
    return path


def safe_batch_tag(value: str | None = None) -> str:
    text = str(value or "").strip()
    if not text:
        text = datetime.now().strftime("%Y%m%d_%H%M%S")
    text = re.sub(r"^WB(\d{8})(\d{6})(?:-([0-9A-Za-z]+))?$", r"\1_\2_\3", text)
    text = re.sub(r"[^0-9A-Za-z_-]+", "_", text).strip("_")
    return text or datetime.now().strftime("%Y%m%d_%H%M%S")


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(2, 1000):
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    return path.with_name(f"{path.stem}_{datetime.now().strftime('%f')}{path.suffix}")


def output_paths(output_dir: Path | None = None, batch_tag: str | None = None) -> tuple[Path, Path]:
    output_dir = output_dir or get_waybill_output_dir()
    base = f"面单信息_{MACHINE_NAME}"
    if batch_tag:
        base = f"{base}_{safe_batch_tag(batch_tag)}"
    return output_dir / f"{base}.xlsx", output_dir / f"{base}.jsonl"


def raw_waybill_path(output_dir: Path | None = None, batch_tag: str | None = None) -> Path:
    output_dir = output_dir or get_waybill_output_dir()
    base = f"监控面单原文_{MACHINE_NAME}"
    if batch_tag:
        base = f"{base}_{safe_batch_tag(batch_tag)}"
    return output_dir / f"{base}.xlsx"


def processed_waybill_path(output_dir: Path | None = None, batch_tag: str | None = None) -> Path:
    output_dir = output_dir or get_waybill_output_dir()
    base = f"监控面单识别_{MACHINE_NAME}"
    if batch_tag:
        base = f"{base}_{safe_batch_tag(batch_tag)}"
    return output_dir / f"{base}.xlsx"


def raw_record_text(record: dict) -> str:
    return str(
        record.get("打印信息")
        or record.get("print_text_raw")
        or record.get("print_text")
        or ""
    ).strip()


def record_key(record: dict) -> str:
    task_id = str(record.get("task_id") or "").strip()
    document_id = str(record.get("document_id") or "").strip()
    if task_id or document_id:
        return "|".join(
            [
                str(record.get("source_client_id") or record.get("machine_name") or "").strip(),
                task_id,
                document_id,
            ]
        )

    source_index = str(record.get("source_record_index") or record.get("record_index") or "").strip()
    if source_index:
        return "|".join(
            [
                str(record.get("source_client_id") or record.get("machine_name") or "").strip(),
                source_index,
                raw_record_text(record),
            ]
        )

    return ""


def record_tracking_values(record: dict) -> dict:
    source_machine = record.get("machine_label") or record.get("machine_name") or ""
    return {
        "任务ID": record.get("task_id", ""),
        "文档ID": record.get("document_id", ""),
        "任务时间": record.get("task_time", ""),
        "采集端ID": record.get("source_client_id", ""),
        "来源机器": source_machine,
        "来源序号": record.get("source_record_index") or record.get("record_index", ""),
    }


def read_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    rows = []
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        if not line.strip():
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return rows


def read_exported_records() -> list[dict]:
    _, jsonl_path = output_paths()
    return read_jsonl(jsonl_path)


def write_jsonl(path: Path, records: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp.jsonl")
    with tmp.open("w", encoding="utf-8", newline="\n") as f:
        for record in records:
            raw_text = raw_record_text(record)
            if raw_text:
                payload = {
                    "打印信息": raw_text,
                    "task_id": record.get("task_id", ""),
                    "document_id": record.get("document_id", ""),
                    "task_time": record.get("task_time", ""),
                    "source_client_id": record.get("source_client_id", ""),
                    "source_record_index": record.get("source_record_index", ""),
                    "machine_name": record.get("machine_name", ""),
                    "machine_label": record.get("machine_label", ""),
                }
                f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    os.replace(tmp, path)


def merge_records(existing: list[dict], incoming: list[dict]) -> tuple[list[dict], int]:
    incoming_by_key = {key: row for row in incoming if (key := record_key(row))}
    merged = []
    seen = set()

    for row in existing:
        key = record_key(row)
        fresh = incoming_by_key.get(key, {}) if key else {}
        merged_row = dict(row)
        for field, value in fresh.items():
            if value and not merged_row.get(field):
                merged_row[field] = value
        merged.append(merged_row)
        if key:
            seen.add(key)

    added = 0
    for row in incoming:
        key = record_key(row)
        if key and key in seen:
            continue
        merged.append(row)
        if key:
            seen.add(key)
        added += 1

    return merged, added


def write_xlsx(path: Path, records: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "面单信息"
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col, (_, label) in enumerate(FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, record in enumerate(records, 2):
        for col, (field, _) in enumerate(FIELDS, 1):
            if field == "print_text":
                value = raw_record_text(record)
            elif field == "source_machine":
                value = record.get("machine_label") or record.get("machine_name") or ""
            else:
                value = record.get(field, "")
            ws.cell(row=row_index, column=col, value=value)
        ws.cell(row=row_index, column=len(FIELDS)).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [24, 24, 20, 28, 22, 12, 100]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    if records:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(FIELDS)).column_letter}{len(records) + 1}"

    tmp = path.with_suffix(".tmp.xlsx")
    try:
        wb.save(tmp)
        os.replace(tmp, path)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass


def export_records(records: list[dict], merge_existing: bool = True, batch_tag: str | None = None) -> dict:
    xlsx_path, jsonl_path = output_paths(batch_tag=batch_tag)
    if batch_tag:
        xlsx_path = unique_path(xlsx_path)
        jsonl_path = xlsx_path.with_suffix(".jsonl")
    if merge_existing:
        existing = read_jsonl(jsonl_path)
        merged, added = merge_records(existing, records)
    else:
        merged = list(records)
        added = len(merged)
    write_xlsx(xlsx_path, merged)
    write_jsonl(jsonl_path, merged)
    return {
        "records_found": len(records),
        "added": added,
        "total": len(merged),
        "xlsx": str(xlsx_path),
        "jsonl": str(jsonl_path),
    }


def build_raw_waybill_rows(records: list[dict]) -> list[dict]:
    rows = []
    for record in records:
        raw_text = raw_record_text(record)
        if not raw_text:
            continue
        row = record_tracking_values(record)
        row[RAW_WAYBILL_TEXT_COLUMN] = raw_text
        rows.append(row)
    return rows


def write_raw_waybill_xlsx(records: list[dict], path: Path | None = None) -> Path:
    path = path or raw_waybill_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = build_raw_waybill_rows(records)

    wb = Workbook()
    ws = wb.active
    ws.title = "打印信息原文"
    headers = RAW_WAYBILL_HEADERS
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for row_index, row in enumerate(rows, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row_index, column=col, value=row.get(header, ""))
    widths = [24, 24, 20, 28, 22, 12, 100]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    if rows:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(rows) + 1}"

    tmp = path.with_suffix(".tmp.xlsx")
    try:
        wb.save(tmp)
        os.replace(tmp, path)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
    return path
