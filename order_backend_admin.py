import os
import sys
import shutil
import base64
import zipfile
import xml.etree.ElementTree as ET
import re
import threading
import queue
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
import openpyxl

from order_secure_common import (
    load_data, save_data, save_templates_fast, get_active_system,
    normalize_text, make_image_key,
    get_data_file, get_data_dir, get_output_dir,
    backup_data_file, preview_data_summary,
    upsert_image_binding, delete_image_binding, clear_all_image_categories,
    iter_image_bindings, image_storage_summary, list_image_category_names,
    ImageMatcher, load_image_map_for_categories
)
from order_core import (
    RuleEngine,
    extract_raw_spec,
    extract_size,
    get_template,
    normalize_qty,
    read_by_template,
)


APP_TITLE = "订单整理管理系统"


def resource_path(filename):
    if getattr(sys, "frozen", False):
        base_dir = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(sys.executable)))
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, filename)


def set_window_icon(root):
    try:
        ico = resource_path("icon_backend.ico")
        if os.path.exists(ico):
            root.iconbitmap(default=ico)
    except Exception:
        pass
    try:
        png = resource_path("icon_backend.png")
        if os.path.exists(png):
            img = tk.PhotoImage(file=png)
            root.iconphoto(True, img)
            root._icon_ref = img
    except Exception:
        pass


def extract_wps_cell_images_from_xlsx(xlsx_path):
    cell_image_map = {}
    with zipfile.ZipFile(xlsx_path, "r") as z:
        names = set(z.namelist())
        if "xl/cellimages.xml" not in names or "xl/_rels/cellimages.xml.rels" not in names:
            return cell_image_map

        rel_root = ET.fromstring(z.read("xl/_rels/cellimages.xml.rels"))
        rel_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
        rid_to_target = {}

        for rel in rel_root.findall("rel:Relationship", rel_ns):
            rid = rel.attrib.get("Id")
            target = rel.attrib.get("Target", "")
            if rid and target and target != "NULL":
                if target.startswith("../"):
                    target = target.replace("../", "xl/")
                elif target.startswith("media/"):
                    target = "xl/" + target
                elif not target.startswith("xl/"):
                    target = "xl/" + target
                rid_to_target[rid] = target

        cell_root = ET.fromstring(z.read("xl/cellimages.xml"))
        ns = {
            "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }

        image_id_to_media = {}
        for cell_img in cell_root.findall(".//etc:cellImage", ns):
            cnv = cell_img.find(".//xdr:cNvPr", ns)
            blip = cell_img.find(".//a:blip", ns)
            if cnv is None or blip is None:
                continue
            image_id = cnv.attrib.get("name")
            rid = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
            media = rid_to_target.get(rid)
            if image_id and media and media in names:
                image_id_to_media[image_id] = media

        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and "DISPIMG" in val:
                        m = re.search(r'DISPIMG\("([^"]+)"', val)
                        if not m:
                            continue
                        media = image_id_to_media.get(m.group(1))
                        if media:
                            cell_image_map[f"{ws.title}!{cell.coordinate}"] = {
                                "bytes": z.read(media),
                                "filename": os.path.basename(media),
                                "sheet": ws.title,
                                "cell": cell.coordinate
                            }
    return cell_image_map


def split_specs_text(text):
    if text is None:
        return []
    return [p.strip() for p in re.split(r"[\n;；]+", str(text).strip()) if p.strip()]


def find_spec_right_of_image(ws, image_cell):
    for offset in range(1, 4):
        value = ws.cell(row=image_cell.row, column=image_cell.column + offset).value
        if value is not None and str(value).strip() and "DISPIMG" not in str(value):
            return str(value).strip()
    return ""



class LoadingWindow:
    def __init__(self, root, on_close=None):
        self.root = root
        self.on_close = on_close
        self.win = tk.Toplevel(root)
        self.win.title("订单整理管理系统 V7.6.1")
        self.win.geometry("760x500")
        self.win.resizable(False, False)
        self.win.protocol("WM_DELETE_WINDOW", self.close_request)

        frame = ttk.Frame(self.win)
        frame.pack(fill="both", expand=True, padx=22, pady=18)

        ttk.Label(
            frame,
            text="正在启动订单整理管理系统",
            font=("Microsoft YaHei", 16, "bold")
        ).pack(anchor="w")

        self.status_var = tk.StringVar(value="准备读取轻量数据")
        self.detail_var = tk.StringVar(value=get_data_file())
        ttk.Label(frame, textvariable=self.status_var, font=("Microsoft YaHei", 10)).pack(anchor="w", pady=(10, 2))
        ttk.Label(frame, textvariable=self.detail_var, foreground="#5f6b7a").pack(anchor="w", pady=(0, 12))

        self.progress_var = tk.DoubleVar(value=0)
        self.progress = ttk.Progressbar(frame, mode="determinate", maximum=100, variable=self.progress_var)
        self.progress.pack(fill="x", pady=(0, 12))

        steps = ttk.LabelFrame(frame, text="加载进度")
        steps.pack(fill="both", expand=True)
        self.text = tk.Text(steps, height=14, width=88, borderwidth=0, bg="#f8fafc")
        self.text.pack(fill="both", expand=True, padx=10, pady=10)

        btns = ttk.Frame(frame)
        btns.pack(fill="x", pady=(14, 0))
        ttk.Button(btns, text="取消并退出", command=self.close_request).pack(side="right")

        self.set_progress(2, "准备读取轻量数据", get_data_file())

    def close_request(self):
        if messagebox.askyesno("确认退出", "数据还在加载，是否强制退出程序？"):
            if self.on_close:
                self.on_close()
            else:
                self.root.destroy()

    def write(self, msg):
        try:
            self.text.insert(tk.END, str(msg) + "\n")
            self.text.see(tk.END)
            self.win.update_idletasks()
        except Exception:
            pass

    def set_progress(self, percent, message, detail=""):
        try:
            percent = max(0, min(100, float(percent)))
            self.progress_var.set(percent)
            self.status_var.set(f"{int(percent)}%  {message}")
            if detail:
                self.detail_var.set(str(detail))
            self.write(f"[{int(percent):>3}%] {message}" + (f" - {detail}" if detail else ""))
        except Exception:
            pass

    def close(self):
        try:
            self.win.destroy()
        except Exception:
            pass


class BackendAdmin:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1180x780")
        self.root.minsize(980, 660)

        self.data = None
        self.system = None
        self.system_id = None
        self.load_queue = queue.Queue()
        self.load_started_at = time.time()
        self.loading_done = False

        self.root.withdraw()
        self.splash = LoadingWindow(self.root, on_close=self.force_exit)
        self.splash.write("后台真实加载线程已启动。")
        self.start_async_load()
        self.poll_load_queue()

    def force_exit(self):
        try:
            self.root.destroy()
        except Exception:
            pass

    def start_async_load(self):
        worker = threading.Thread(target=self._load_data_worker, daemon=True)
        worker.start()

    def _load_data_worker(self):
        try:
            def progress(percent, message, detail=""):
                self.load_queue.put(("progress", percent, message, detail))

            progress(5, "检查运行目录", get_data_dir())
            data = load_data(progress=progress)

            progress(94, "扫描图片数据目录", "只统计文件，不加载全部图片关系")
            image_stats = image_storage_summary(count_entries=False)
            progress(
                96,
                "图片数据扫描完成",
                f"{image_stats.get('category_files', 0)} 个分类文件，"
                f"{image_stats.get('image_files', 0)} 个图片文件，"
                f"{image_stats.get('bytes', 0) / 1024 / 1024:.2f} MB"
            )

            progress(98, "统计模板、规则、档口", "准备初始化主界面")
            summary = preview_data_summary(data)

            progress(100, "加载完成", "正在打开主界面")
            self.load_queue.put(("done", data, summary))
        except Exception as e:
            self.load_queue.put(("error", e))

    def poll_load_queue(self):
        try:
            while True:
                msg = self.load_queue.get_nowait()
                kind = msg[0]

                if kind == "log":
                    self.splash.write(msg[1])

                elif kind == "progress":
                    _, percent, message, detail = msg
                    self.splash.set_progress(percent, message, detail)

                elif kind == "error":
                    err = msg[1]
                    self.splash.write(f"加载失败：{err}")
                    messagebox.showerror("加载失败", str(err))
                    self.force_exit()
                    return

                elif kind == "done":
                    _, data, summary = msg
                    self.finish_loading(data, summary)
                    return
        except queue.Empty:
            pass

        self.root.after(200, self.poll_load_queue)

    def finish_loading(self, data, summary):
        self.loading_done = True
        self.data = data
        self.system, self.system_id = get_active_system(self.data)

        try:
            self.splash.write("阶段 5/5：初始化主界面")
            self.splash.write(f"当前系统：{summary.get('system_name', '')}")
            self.splash.write(f"模板数量：{summary.get('templates_count', 0)}")
            self.splash.write(f"分类规则：{summary.get('rules_count', 0)}")
            self.splash.write(f"档口规则：{summary.get('stalls_count', 0)}")
            self.splash.write(
                f"图片数据：{summary.get('image_category_files', 0)} 个分类文件，"
                f"{summary.get('images_count', 0)} 条关系，{summary.get('image_storage_mb', 0)} MB"
            )
            if summary.get("templates_preview"):
                self.splash.write("模板预览：" + "，".join(summary.get("templates_preview", [])))
            if summary.get("rules_preview"):
                self.splash.write("规则预览：" + "，".join(summary.get("rules_preview", [])))
        except Exception:
            pass

        self.template_name = tk.StringVar()
        self.template_mode = tk.StringVar(value="表头")
        self.template_short = tk.StringVar(value="商品简称")
        self.template_spec = tk.StringVar(value="销售规格")
        self.template_qty = tk.StringVar(value="商品数量")
        self.template_remark = tk.StringVar(value="备注")
        self.template_title_col = tk.StringVar(value="S")
        self.template_qty_col = tk.StringVar(value="V")
        self.template_item_sep = tk.StringVar(value=";")
        self.template_spec_split = tk.StringVar(value="，")
        self.active_template = tk.StringVar(value="")

        # 样本Excel字段识别
        self.sample_headers = []
        self.sample_short_header = tk.StringVar()
        self.sample_spec_header = tk.StringVar()
        self.sample_qty_header = tk.StringVar()
        self.sample_remark_header = tk.StringVar()

        self.rule_category = tk.StringVar()
        self.rule_keyword = tk.StringVar()
        self.rule_field = tk.StringVar(value="商品简称")
        self.rule_remove = tk.StringVar()

        self.stall_category = tk.StringVar()
        self.stall_name = tk.StringVar()

        self.img_category = tk.StringVar()
        self.img_spec = tk.StringVar()
        self.img_path = tk.StringVar()
        self.image_search_var = tk.StringVar()
        self.image_filter_category_var = tk.StringVar(value="全部分类")
        self.order_review_rows = {}
        self.order_review_files = []
        self.order_review_template = tk.StringVar()
        self.order_review_category = tk.StringVar()
        self.order_review_keyword = tk.StringVar()
        self.order_review_field = tk.StringVar(value="商品简称")
        self.order_review_remove = tk.StringVar()
        self.order_review_search = tk.StringVar()
        self.order_review_show_known = tk.BooleanVar(value=True)

        self.build_ui()
        self.refresh_all()

        try:
            self.splash.close()
        except Exception:
            pass

        self.root.deiconify()

    def save_all(self):
        self.data["systems"][self.system_id] = self.system
        save_data(self.data)

    def build_ui(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        nb = ttk.Notebook(self.root)
        nb.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        self.tab_templates = ttk.Frame(nb)
        self.tab_order_review = ttk.Frame(nb)
        self.tab_rules = ttk.Frame(nb)
        self.tab_stalls = ttk.Frame(nb)
        self.tab_images = ttk.Frame(nb)
        self.tab_system = ttk.Frame(nb)

        nb.add(self.tab_templates, text="导入模板")
        nb.add(self.tab_order_review, text="订单识别")
        nb.add(self.tab_rules, text="分类规则")
        nb.add(self.tab_stalls, text="档口规则")
        nb.add(self.tab_images, text="图片关系")
        nb.add(self.tab_system, text="系统维护")

        self.build_templates()
        self.build_order_review()
        self.build_rules()
        self.build_stalls()
        self.build_images()
        self.build_system()

    def build_templates(self):
        self.tab_templates.columnconfigure(0, weight=1)
        self.tab_templates.rowconfigure(2, weight=1)

        top = ttk.LabelFrame(self.tab_templates, text="导入模板配置")
        top.grid(row=0, column=0, sticky="ew", padx=12, pady=10)

        fields = [
            ("模板名称", self.template_name), ("模式", self.template_mode),
            ("商品简称表头", self.template_short), ("规格表头", self.template_spec),
            ("数量表头", self.template_qty), ("备注表头", self.template_remark),
            ("标题列", self.template_title_col), ("数量列", self.template_qty_col), ("商品分隔符", self.template_item_sep),
            ("规格拆分符", self.template_spec_split),
        ]

        for i, (label, var) in enumerate(fields):
            row = i // 3
            col = (i % 3) * 2
            ttk.Label(top, text=label).grid(row=row, column=col, padx=5, pady=6)

            if label == "模式":
                ttk.Combobox(
                    top,
                    textvariable=var,
                    values=["表头", "列号"],
                    state="readonly",
                    width=16
                ).grid(row=row, column=col + 1, padx=5)
            else:
                ttk.Entry(top, textvariable=var, width=22).grid(row=row, column=col + 1, padx=5)

        ttk.Button(
            top,
            text="保存模板",
            command=self.save_template
        ).grid(row=4, column=0, padx=5, pady=8)

        ttk.Button(
            top,
            text="删除选中模板",
            command=self.delete_template
        ).grid(row=4, column=1, padx=5, pady=8)
        ttk.Button(top, text="读取样本表头", command=self.preview_excel_headers).grid(row=4, column=2, padx=5)

        ttk.Label(top, text="模板说明：").grid(row=5, column=0, padx=5, pady=6)
        ttk.Label(
            top,
            text="备注表头为可选项；填写后分类识别和图片匹配会同时参考订单备注。",
            foreground="#555555"
        ).grid(row=5, column=1, columnspan=4, sticky="w", padx=5)

        # =========================
        # 自定义识别导入数据
        # =========================
        custom_box = ttk.LabelFrame(self.tab_templates, text="自定义识别导入数据")
        custom_box.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 10))

        ttk.Label(
            custom_box,
            text="选择一个1688导出的样本Excel，系统读取表头后，你可以指定商品简称、销售规格、数量和备注字段。"
        ).grid(row=0, column=0, columnspan=8, sticky="w", padx=8, pady=8)

        ttk.Button(
            custom_box,
            text="选择样本Excel并识别字段",
            command=self.load_sample_excel_headers
        ).grid(row=1, column=0, padx=8, pady=8)

        ttk.Label(custom_box, text="商品简称字段").grid(row=1, column=1, padx=5)
        self.sample_short_combo = ttk.Combobox(
            custom_box,
            textvariable=self.sample_short_header,
            values=[],
            width=22,
            state="readonly"
        )
        self.sample_short_combo.grid(row=1, column=2, padx=5)

        ttk.Label(custom_box, text="销售规格字段").grid(row=1, column=3, padx=5)
        self.sample_spec_combo = ttk.Combobox(
            custom_box,
            textvariable=self.sample_spec_header,
            values=[],
            width=22,
            state="readonly"
        )
        self.sample_spec_combo.grid(row=1, column=4, padx=5)

        ttk.Label(custom_box, text="数量字段").grid(row=1, column=5, padx=5)
        self.sample_qty_combo = ttk.Combobox(
            custom_box,
            textvariable=self.sample_qty_header,
            values=[],
            width=18,
            state="readonly"
        )
        self.sample_qty_combo.grid(row=1, column=6, padx=5)

        ttk.Label(custom_box, text="备注字段").grid(row=2, column=1, padx=5)
        self.sample_remark_combo = ttk.Combobox(
            custom_box,
            textvariable=self.sample_remark_header,
            values=[],
            width=22,
            state="readonly"
        )
        self.sample_remark_combo.grid(row=2, column=2, padx=5, pady=5)

        ttk.Button(
            custom_box,
            text="生成并保存为模板",
            command=self.save_template_from_sample
        ).grid(row=2, column=3, padx=8)

        ttk.Label(
            custom_box,
            text="说明：这个功能用于1688导出格式变化时重新配置字段，配置后前端无需改程序。"
        ).grid(row=3, column=0, columnspan=8, sticky="w", padx=8, pady=(0, 8))

        cols = ("模板名称", "模式", "商品简称", "规格", "数量", "备注", "标题列", "数量列", "商品分隔符", "规格拆分符")
        self.template_tree = ttk.Treeview(self.tab_templates, columns=cols, show="headings")

        for c in cols:
            self.template_tree.heading(c, text=c)
            self.template_tree.column(c, width=120, anchor="center")

        self.template_tree.grid(row=2, column=0, sticky="nsew", padx=12, pady=8)
        self.template_tree.bind("<Double-1>", self.load_template)

    def build_order_review(self):
        self.tab_order_review.columnconfigure(0, weight=1)
        self.tab_order_review.rowconfigure(2, weight=1)

        top = ttk.LabelFrame(self.tab_order_review, text="从订单识别并补分类规则")
        top.grid(row=0, column=0, sticky="ew", padx=12, pady=10)
        top.columnconfigure(8, weight=1)

        ttk.Label(
            top,
            text="保留原整理模式：先读取1688订单，查看当前识别结果，编辑后保存为分类规则。"
        ).grid(row=0, column=0, columnspan=9, sticky="w", padx=8, pady=8)

        ttk.Button(top, text="导入订单Excel", command=self.import_order_review_excel).grid(row=1, column=0, padx=8, pady=8)
        ttk.Button(top, text="重新识别", command=self.reload_order_review).grid(row=1, column=1, padx=5)
        ttk.Label(top, text="导入模板").grid(row=1, column=2, padx=5)
        self.order_review_template_combo = ttk.Combobox(
            top,
            textvariable=self.order_review_template,
            width=26,
            state="readonly"
        )
        self.order_review_template_combo.grid(row=1, column=3, padx=5)
        ttk.Button(top, text="保存选中规则", command=self.save_selected_order_review_rules).grid(row=1, column=4, padx=5)
        ttk.Button(top, text="批量保存可确认", command=self.save_all_ready_order_review_rules).grid(row=1, column=5, padx=5)
        ttk.Checkbutton(
            top,
            text="显示已识别",
            variable=self.order_review_show_known,
            command=self.refresh_order_review
        ).grid(row=1, column=6, padx=5, sticky="w")

        edit = ttk.LabelFrame(self.tab_order_review, text="编辑选中行要保存的规则")
        edit.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 8))
        edit.columnconfigure(7, weight=1)

        ttk.Label(edit, text="分类/鞋款").grid(row=0, column=0, padx=5, pady=8)
        ttk.Entry(edit, textvariable=self.order_review_category, width=22).grid(row=0, column=1, padx=5)
        ttk.Label(edit, text="关键词").grid(row=0, column=2, padx=5)
        ttk.Entry(edit, textvariable=self.order_review_keyword, width=26).grid(row=0, column=3, padx=5)
        ttk.Label(edit, text="匹配位置").grid(row=0, column=4, padx=5)
        ttk.Combobox(
            edit,
            textvariable=self.order_review_field,
            values=["商品简称", "销售规格", "备注", "货品标题", "全部"],
            state="readonly",
            width=12
        ).grid(row=0, column=5, padx=5)
        ttk.Button(edit, text="关键词=商品简称", command=self.use_order_review_short_keyword).grid(row=0, column=6, padx=5)
        ttk.Button(edit, text="关键词=销售规格", command=self.use_order_review_spec_keyword).grid(row=0, column=7, padx=5, sticky="w")

        ttk.Label(edit, text="清洗关键词").grid(row=1, column=0, padx=5, pady=8)
        ttk.Entry(edit, textvariable=self.order_review_remove, width=40).grid(row=1, column=1, columnspan=3, sticky="ew", padx=5)
        ttk.Label(edit, text="搜索").grid(row=1, column=4, padx=5)
        ttk.Entry(edit, textvariable=self.order_review_search).grid(row=1, column=5, columnspan=3, sticky="ew", padx=5)
        self.order_review_search.trace_add("write", lambda *args: self.refresh_order_review())

        table_frame = ttk.Frame(self.tab_order_review)
        table_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=8)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        cols = ("状态", "商品简称", "当前分类", "销售规格", "尺码", "数量", "备注", "图片")
        self.order_review_tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="extended")
        widths = {
            "状态": 80,
            "商品简称": 170,
            "当前分类": 130,
            "销售规格": 260,
            "尺码": 70,
            "数量": 70,
            "备注": 220,
            "图片": 80,
        }
        for col in cols:
            self.order_review_tree.heading(col, text=col)
            self.order_review_tree.column(col, width=widths.get(col, 120), anchor="center" if col in {"状态", "尺码", "数量", "图片"} else "w")
        self.order_review_tree.grid(row=0, column=0, sticky="nsew")
        self.order_review_tree.bind("<<TreeviewSelect>>", self.load_order_review_selection)
        self.order_review_tree.bind("<Double-1>", self.load_order_review_selection)

        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.order_review_tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        self.order_review_tree.configure(yscrollcommand=y_scroll.set)

        bottom = ttk.Frame(self.tab_order_review)
        bottom.grid(row=3, column=0, sticky="ew", padx=12, pady=(0, 10))
        self.order_review_count_label = ttk.Label(bottom, text="未导入订单")
        self.order_review_count_label.pack(side="left", padx=5)

    def build_rules(self):
        self.tab_rules.columnconfigure(0, weight=1)
        self.tab_rules.rowconfigure(1, weight=1)

        top = ttk.LabelFrame(self.tab_rules, text="分类识别规则")
        top.grid(row=0, column=0, sticky="ew", padx=12, pady=10)

        ttk.Label(top, text="分类").grid(row=0, column=0, padx=5, pady=8)
        ttk.Entry(top, textvariable=self.rule_category, width=20).grid(row=0, column=1, padx=5)
        ttk.Label(top, text="关键词").grid(row=0, column=2, padx=5)
        ttk.Entry(top, textvariable=self.rule_keyword, width=24).grid(row=0, column=3, padx=5)
        ttk.Label(top, text="匹配位置").grid(row=0, column=4, padx=5)
        ttk.Combobox(top, textvariable=self.rule_field, values=["商品简称", "销售规格", "备注", "货品标题", "全部"], state="readonly", width=12).grid(row=0, column=5, padx=5)

        ttk.Label(top, text="清洗关键词").grid(row=1, column=0, padx=5, pady=8)
        ttk.Entry(top, textvariable=self.rule_remove, width=52).grid(row=1, column=1, columnspan=3, padx=5, sticky="ew")
        ttk.Button(top, text="保存规则", command=self.save_rule).grid(row=1, column=4, padx=5)
        ttk.Button(top, text="删除选中", command=self.delete_rule).grid(row=1, column=5, padx=5)

        cols = ("分类", "关键词", "匹配位置", "清洗关键词")
        self.rule_tree = ttk.Treeview(self.tab_rules, columns=cols, show="headings")
        for c in cols:
            self.rule_tree.heading(c, text=c)
            self.rule_tree.column(c, width=180 if c != "清洗关键词" else 420, anchor="center" if c != "清洗关键词" else "w")
        self.rule_tree.grid(row=1, column=0, sticky="nsew", padx=12, pady=8)
        self.rule_tree.bind("<Double-1>", self.load_rule)

    def build_stalls(self):
        self.tab_stalls.columnconfigure(0, weight=1)
        self.tab_stalls.rowconfigure(1, weight=1)

        top = ttk.LabelFrame(self.tab_stalls, text="档口规则")
        top.grid(row=0, column=0, sticky="ew", padx=12, pady=10)

        ttk.Label(top, text="分类").grid(row=0, column=0, padx=5, pady=8)
        self.stall_category_combo = ttk.Combobox(top, textvariable=self.stall_category, width=24)
        self.stall_category_combo.grid(row=0, column=1, padx=5)

        ttk.Label(top, text="档口").grid(row=0, column=2, padx=5)
        ttk.Entry(top, textvariable=self.stall_name, width=28).grid(row=0, column=3, padx=5)

        ttk.Button(top, text="保存档口规则", command=self.save_stall).grid(row=0, column=4, padx=5)
        ttk.Button(top, text="删除选中", command=self.delete_stall).grid(row=0, column=5, padx=5)

        cols = ("分类", "档口")
        self.stall_tree = ttk.Treeview(self.tab_stalls, columns=cols, show="headings")
        for c in cols:
            self.stall_tree.heading(c, text=c)
            self.stall_tree.column(c, width=320, anchor="center")
        self.stall_tree.grid(row=1, column=0, sticky="nsew", padx=12, pady=8)
        self.stall_tree.bind("<Double-1>", self.load_stall)

    def build_images(self):
        self.tab_images.columnconfigure(0, weight=1)
        self.tab_images.rowconfigure(2, weight=1)

        top = ttk.LabelFrame(self.tab_images, text="图片关系：分类 + 规格 = 图片")
        top.grid(row=0, column=0, sticky="ew", padx=12, pady=10)
        top.columnconfigure(6, weight=1)

        ttk.Label(top, text="分类").grid(row=0, column=0, padx=5, pady=8)
        self.img_category_combo = ttk.Combobox(top, textvariable=self.img_category, width=24)
        self.img_category_combo.grid(row=0, column=1, padx=5)

        ttk.Label(top, text="规格").grid(row=0, column=2, padx=5)
        ttk.Entry(top, textvariable=self.img_spec, width=30).grid(row=0, column=3, padx=5)

        ttk.Button(top, text="选择图片", command=self.choose_image).grid(row=0, column=4, padx=5)
        ttk.Button(top, text="保存图片", command=self.save_image).grid(row=0, column=5, padx=5)
        ttk.Button(top, text="从Excel批量导入", command=self.import_images_from_excel).grid(row=0, column=6, padx=5, sticky="w")

        ttk.Entry(top, textvariable=self.img_path).grid(row=1, column=0, columnspan=7, sticky="ew", padx=5, pady=8)

        filter_box = ttk.LabelFrame(self.tab_images, text="搜索 / 筛选")
        filter_box.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 8))
        filter_box.columnconfigure(3, weight=1)

        ttk.Label(filter_box, text="分类筛选").grid(row=0, column=0, padx=5, pady=8)
        self.image_filter_category_combo = ttk.Combobox(filter_box, textvariable=self.image_filter_category_var, values=["全部分类"], width=24, state="readonly")
        self.image_filter_category_combo.grid(row=0, column=1, padx=5, pady=8)

        ttk.Label(filter_box, text="关键词搜索").grid(row=0, column=2, padx=5, pady=8)
        ttk.Entry(filter_box, textvariable=self.image_search_var).grid(row=0, column=3, sticky="ew", padx=5, pady=8)

        ttk.Button(filter_box, text="搜索", command=self.refresh_images).grid(row=0, column=4, padx=5)
        ttk.Button(filter_box, text="重置", command=self.reset_image_filter).grid(row=0, column=5, padx=5)

        self.image_search_var.trace_add("write", lambda *args: self.refresh_images())
        self.image_filter_category_var.trace_add("write", lambda *args: self.refresh_images())

        table_frame = ttk.Frame(self.tab_images)
        table_frame.grid(row=2, column=0, sticky="nsew", padx=12, pady=8)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        cols = ("分类", "规格", "状态")
        self.image_tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="extended")
        for c in cols:
            self.image_tree.heading(c, text=c)
        self.image_tree.column("分类", width=180, anchor="center")
        self.image_tree.column("规格", width=420, anchor="center")
        self.image_tree.column("状态", width=160, anchor="center")
        self.image_tree.grid(row=0, column=0, sticky="nsew")

        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.image_tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        self.image_tree.configure(yscrollcommand=y_scroll.set)

        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.image_tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        self.image_tree.configure(xscrollcommand=x_scroll.set)

        bottom = ttk.Frame(self.tab_images)
        bottom.grid(row=3, column=0, sticky="ew", padx=12, pady=8)
        ttk.Button(bottom, text="删除选中图片绑定", command=self.delete_selected_images).pack(side="left", padx=5)
        ttk.Button(bottom, text="删除当前筛选结果", command=self.delete_filtered_images).pack(side="left", padx=5)
        ttk.Button(bottom, text="清空全部图片关系", command=self.clear_all_images).pack(side="left", padx=5)

        self.image_count_label = ttk.Label(bottom, text="当前显示：0 条")
        self.image_count_label.pack(side="right", padx=5)

    def build_system(self):
        box = ttk.LabelFrame(self.tab_system, text="系统维护")
        box.pack(fill="x", padx=12, pady=12)

        ttk.Label(box, text=f"当前整理系统：{self.system.get('name', self.system_id)}").pack(anchor="w", padx=10, pady=8)

        ttk.Button(box, text="导出配置包", command=self.export_config).pack(side="left", padx=10, pady=10)
        ttk.Button(box, text="导入配置包", command=self.import_config).pack(side="left", padx=10, pady=10)
        ttk.Button(box, text="备份加密数据", command=self.backup_data).pack(side="left", padx=10, pady=10)
        ttk.Button(box, text="重建轻量主数据", command=self.repair_old_data).pack(side="left", padx=10, pady=10)
        ttk.Button(box, text="查看主数据大小", command=self.compact_data).pack(side="left", padx=10, pady=10)
        ttk.Button(box, text="打开数据目录", command=lambda: os.startfile(get_data_dir())).pack(side="left", padx=10, pady=10)
        ttk.Button(box, text="打开输出目录", command=lambda: os.startfile(get_output_dir())).pack(side="left", padx=10, pady=10)

        preview_box = ttk.LabelFrame(self.tab_system, text="当前数据预览")
        preview_box.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        self.data_preview_text = tk.Text(preview_box, height=12)
        self.data_preview_text.pack(fill="both", expand=True, padx=10, pady=10)
        self.refresh_data_preview()

    def refresh_data_preview(self):
        if not hasattr(self, "data_preview_text"):
            return

        try:
            summary = preview_data_summary(self.data)
            lines = [
                f"数据文件：{get_data_file()}",
                f"当前系统：{summary.get('system_name', '')}",
                f"模板数量：{summary.get('templates_count', 0)}",
                f"分类规则：{summary.get('rules_count', 0)}",
                f"档口规则：{summary.get('stalls_count', 0)}",
                f"图片分片：{summary.get('image_category_files', 0)} 个分类文件",
                f"图片关系：{summary.get('images_count', 0)} 条",
                f"图片数据大小：{summary.get('image_storage_mb', 0)} MB",
                "",
                "模板预览：",
            ]
            lines.extend([f"  - {x}" for x in summary.get("templates_preview", [])])
            lines.append("")
            lines.append("规则预览：")
            lines.extend([f"  - {x}" for x in summary.get("rules_preview", [])])
            lines.append("")
            lines.append("图片分类预览：")
            lines.extend([f"  - {x}" for x in summary.get("images_preview", [])])

            self.data_preview_text.delete("1.0", tk.END)
            self.data_preview_text.insert(tk.END, "\n".join(lines))
        except Exception as e:
            self.data_preview_text.delete("1.0", tk.END)
            self.data_preview_text.insert(tk.END, f"数据预览失败：{e}")

    def repair_old_data(self):
        if not messagebox.askyesno(
            "确认",
            "即将备份并重建轻量主数据：\n\n"
            "1. 主数据只保留模板、分类规则、档口规则和用户\n"
            "2. 图片关系继续保存在 data/image_categories 分类文件中\n"
            "3. 不再把图片关系写入 system_data.enc\n\n"
            "是否继续？"
        ):
            return

        try:
            backup = backup_data_file()
            save_data(self.data)

            self.data = load_data(auto_save_on_read=False)
            self.system, self.system_id = get_active_system(self.data)
            self.refresh_all()
            self.refresh_data_preview()
            summary = preview_data_summary(self.data)

            messagebox.showinfo(
                "重建完成",
                "轻量主数据已重建。\n\n"
                f"备份文件：{backup or '无旧文件'}\n"
                f"模板数量：{summary.get('templates_count', 0)}\n"
                f"分类规则：{summary.get('rules_count', 0)}\n"
                f"档口规则：{summary.get('stalls_count', 0)}\n"
                f"图片分片：{summary.get('image_category_files', 0)} 个文件\n"
                f"图片关系：{summary.get('images_count', 0)} 条"
            )
        except Exception as e:
            messagebox.showerror("重建失败", str(e))

    def refresh_all(self):
        self.refresh_templates()
        self.refresh_order_review_template_choices()
        self.refresh_order_review()
        self.refresh_rules()
        self.refresh_stalls()
        self.refresh_categories()
        self.refresh_images()

    def refresh_categories(self):
        cats = sorted(
            {r.get("category", "") for r in self.system.get("category_rules", []) if r.get("category", "")}
            | set(list_image_category_names())
        )
        if hasattr(self, "stall_category_combo"):
            self.stall_category_combo["values"] = cats
        if hasattr(self, "img_category_combo"):
            self.img_category_combo["values"] = cats
        if hasattr(self, "image_filter_category_combo"):
            values = ["全部分类"] + cats
            current = self.image_filter_category_var.get()
            self.image_filter_category_combo["values"] = values
            if current not in values:
                self.image_filter_category_var.set("全部分类")

    def refresh_order_review_template_choices(self):
        if not hasattr(self, "order_review_template_combo"):
            return
        names = [t.get("name", "") for t in self.system.get("import_templates", []) if t.get("name", "")]
        self.order_review_template_combo["values"] = names
        current = self.order_review_template.get()
        if current not in names:
            self.order_review_template.set(names[0] if names else "")

    def current_order_review_template(self):
        name = self.order_review_template.get().strip()
        if not name:
            self.refresh_order_review_template_choices()
            name = self.order_review_template.get().strip()
        return get_template(self.system, name)

    def import_order_review_excel(self):
        files = filedialog.askopenfilenames(title="选择1688订单Excel", filetypes=[("Excel文件", "*.xlsx *.xls")])
        if not files:
            return
        self.order_review_files = list(files)
        self.load_order_review_files()

    def reload_order_review(self):
        if not self.order_review_files:
            messagebox.showwarning("提示", "请先导入订单Excel")
            return
        self.load_order_review_files()

    def load_order_review_files(self):
        try:
            template = self.current_order_review_template()
            rule_engine = RuleEngine(self.system.get("category_rules", []))
            groups = {}
            for file_path in self.order_review_files:
                df = read_by_template(file_path, template)
                for _, row in df.iterrows():
                    short_name = str(row.get("商品简称", "") or "").strip()
                    title = str(row.get("货品标题", "") or "").strip()
                    remark = str(row.get("备注", "") or "").strip()
                    raw_spec = extract_raw_spec(title)
                    if raw_spec == "未知":
                        raw_spec = ""
                    size = extract_size(title)
                    category = rule_engine.detect_category(short_name, title, raw_spec, remark)
                    clean_spec = rule_engine.clean_spec(raw_spec, category) if category != "未分类" else normalize_text(raw_spec)
                    key = "|".join([
                        normalize_text(short_name),
                        normalize_text(raw_spec),
                        normalize_text(size),
                        normalize_text(remark),
                    ])
                    if not key.strip("|"):
                        continue
                    item = groups.setdefault(key, {
                        "key": key,
                        "short_name": short_name,
                        "title": title,
                        "raw_spec": raw_spec,
                        "clean_spec": clean_spec,
                        "size": size,
                        "remark": remark,
                        "category": category,
                        "quantity": 0,
                        "image_status": "-",
                    })
                    item["quantity"] += normalize_qty(row.get("数量", 1))

            used_categories = sorted({normalize_text(v.get("category", "")) for v in groups.values() if normalize_text(v.get("category", "")) and v.get("category") != "未分类"})
            image_map = load_image_map_for_categories(self.system, used_categories)
            matcher = ImageMatcher(image_map)
            for item in groups.values():
                category = item.get("category", "")
                if not category or category == "未分类":
                    item["image_status"] = "-"
                    continue
                found = matcher.find(
                    category,
                    item.get("clean_spec", "") or item.get("raw_spec", ""),
                    item.get("remark", ""),
                    item.get("title", ""),
                    item.get("short_name", "")
                )
                item["image_status"] = "已匹配" if found else "无图片"

            self.order_review_rows = groups
            self.refresh_order_review()
            pending = sum(1 for v in groups.values() if v.get("category") == "未分类")
            messagebox.showinfo("读取完成", f"读取到 {len(groups)} 条订单规格；待编辑 {pending} 条。")
        except Exception as e:
            messagebox.showerror("读取失败", str(e))

    def order_review_display_rows(self):
        keyword = normalize_text(self.order_review_search.get())
        rows = []
        for row in self.order_review_rows.values():
            status = "待编辑" if row.get("category") == "未分类" else "已识别"
            if not self.order_review_show_known.get() and status == "已识别":
                continue
            haystack = normalize_text(" ".join([
                row.get("short_name", ""),
                row.get("category", ""),
                row.get("raw_spec", ""),
                row.get("remark", ""),
                row.get("title", ""),
            ]))
            if keyword and keyword not in haystack:
                continue
            item = dict(row)
            item["status"] = status
            rows.append(item)
        order = {"待编辑": 0, "已识别": 1}
        return sorted(rows, key=lambda x: (order.get(x.get("status"), 9), x.get("short_name", ""), x.get("raw_spec", ""), x.get("size", "")))

    def refresh_order_review(self):
        if not hasattr(self, "order_review_tree"):
            return
        for item in self.order_review_tree.get_children():
            self.order_review_tree.delete(item)
        rows = self.order_review_display_rows()
        for row in rows:
            self.order_review_tree.insert(
                "",
                "end",
                iid=row.get("key"),
                values=(
                    row.get("status", ""),
                    row.get("short_name", ""),
                    row.get("category", ""),
                    row.get("raw_spec", ""),
                    row.get("size", ""),
                    row.get("quantity", ""),
                    row.get("remark", ""),
                    row.get("image_status", ""),
                )
            )
        total = len(self.order_review_rows)
        pending = sum(1 for row in self.order_review_rows.values() if row.get("category") == "未分类")
        selected = len(self.order_review_tree.selection())
        self.order_review_count_label.config(text=f"显示：{len(rows)} / 总数：{total} / 待编辑：{pending} / 已选：{selected}")

    def current_order_review_row(self):
        if not hasattr(self, "order_review_tree"):
            return {}
        selected = list(self.order_review_tree.selection())
        key = selected[0] if selected else self.order_review_tree.focus()
        return dict(self.order_review_rows.get(key, {})) if key else {}

    def load_order_review_selection(self, event=None):
        row = self.current_order_review_row()
        if not row:
            return
        category = "" if row.get("category") == "未分类" else row.get("category", "")
        self.order_review_category.set(category)
        self.order_review_keyword.set(row.get("short_name", ""))
        self.order_review_field.set("商品简称")
        selected = len(self.order_review_tree.selection()) if hasattr(self, "order_review_tree") else 0
        total = len(self.order_review_rows)
        pending = sum(1 for item in self.order_review_rows.values() if item.get("category") == "未分类")
        self.order_review_count_label.config(text=f"显示：{len(self.order_review_tree.get_children())} / 总数：{total} / 待编辑：{pending} / 已选：{selected}")

    def use_order_review_short_keyword(self):
        row = self.current_order_review_row()
        if row:
            self.order_review_keyword.set(row.get("short_name", ""))
            self.order_review_field.set("商品简称")

    def use_order_review_spec_keyword(self):
        row = self.current_order_review_row()
        if row:
            self.order_review_keyword.set(row.get("raw_spec", ""))
            self.order_review_field.set("销售规格")

    def order_review_keyword_for_row(self, row, field, override="", single=False):
        if single and override:
            return normalize_text(override)
        if field == "销售规格":
            return normalize_text(row.get("raw_spec", ""))
        if field == "备注":
            return normalize_text(row.get("remark", ""))
        if field == "货品标题":
            return normalize_text(row.get("title", ""))
        if field == "全部" and override:
            return normalize_text(override)
        return normalize_text(row.get("short_name", ""))

    def upsert_category_rule(self, category, keyword, field, remove_words=""):
        category = normalize_text(category)
        keyword = normalize_text(keyword)
        field = field or "商品简称"
        if not category or not keyword:
            return False
        rule = {
            "category": category,
            "keyword": keyword,
            "field": field,
            "remove_words": str(remove_words or "").strip(),
        }
        arr = self.system.setdefault("category_rules", [])
        for i, old in enumerate(arr):
            if (
                normalize_text(old.get("category", "")) == category
                and normalize_text(old.get("keyword", "")) == keyword
                and (old.get("field", "全部") or "全部") == field
            ):
                arr[i] = rule
                return True
        arr.append(rule)
        return True

    def save_selected_order_review_rules(self):
        selected = list(self.order_review_tree.selection()) if hasattr(self, "order_review_tree") else []
        if not selected:
            messagebox.showwarning("提示", "请先选择订单行")
            return
        category = self.order_review_category.get().strip()
        field = self.order_review_field.get() or "商品简称"
        override = self.order_review_keyword.get().strip()
        remove_words = self.order_review_remove.get().strip()
        single = len(selected) == 1
        count = 0
        skipped = 0
        for key in selected:
            row = self.order_review_rows.get(key)
            if not row:
                skipped += 1
                continue
            save_category = category or ("" if row.get("category") == "未分类" else row.get("category", ""))
            keyword = self.order_review_keyword_for_row(row, field, override, single=single)
            if self.upsert_category_rule(save_category, keyword, field, remove_words):
                row["category"] = normalize_text(save_category)
                count += 1
            else:
                skipped += 1
        if count:
            self.save_all()
            self.refresh_all()
            messagebox.showinfo("成功", f"已保存 {count} 条分类规则，跳过 {skipped} 条。")
        else:
            messagebox.showwarning("提示", "没有可保存的规则，请填写分类和关键词。")

    def save_all_ready_order_review_rules(self):
        if not self.order_review_rows:
            messagebox.showwarning("提示", "请先导入订单Excel")
            return
        category = self.order_review_category.get().strip()
        field = self.order_review_field.get() or "商品简称"
        remove_words = self.order_review_remove.get().strip()
        count = 0
        skipped = 0
        for row in self.order_review_rows.values():
            save_category = category or ("" if row.get("category") == "未分类" else row.get("category", ""))
            keyword = self.order_review_keyword_for_row(row, field)
            if self.upsert_category_rule(save_category, keyword, field, remove_words):
                row["category"] = normalize_text(save_category)
                count += 1
            else:
                skipped += 1
        if count:
            self.save_all()
            self.refresh_all()
        messagebox.showinfo("完成", f"已保存 {count} 条规则，跳过 {skipped} 条。")

    def refresh_templates(self):
        if not hasattr(self, "template_tree"):
            return

        for i in self.template_tree.get_children():
            self.template_tree.delete(i)

        names = []
        for t in self.system.get("import_templates", []):
            name = t.get("name", "")
            if name:
                names.append(name)
            self.template_tree.insert(
                "",
                "end",
                values=(
                    t.get("name", ""),
                    t.get("mode", ""),
                    t.get("short_name", ""),
                    t.get("spec", ""),
                    t.get("qty", ""),
                    t.get("remark", ""),
                    t.get("title_col", ""),
                    t.get("qty_col", ""),
                    t.get("item_sep", ""),
                    t.get("spec_split", "")
                )
            )

        if hasattr(self, "active_combo"):
            self.active_combo["values"] = names
            current = self.active_template.get()
            if current not in names:
                self.active_template.set(names[0] if names else "")

    def load_sample_excel_headers(self):
        """
        选择样本Excel并读取表头，用于自定义导入数据识别。
        """
        path = filedialog.askopenfilename(
            title="选择1688导出的样本Excel",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )

        if not path:
            return

        try:
            df = pd.read_excel(path, nrows=5)
            headers = [str(c).strip() for c in df.columns if str(c).strip()]

            if not headers:
                messagebox.showwarning("提示", "没有读取到有效表头")
                return

            self.sample_headers = headers

            self.sample_short_combo["values"] = headers
            self.sample_spec_combo["values"] = headers
            self.sample_qty_combo["values"] = headers
            self.sample_remark_combo["values"] = [""] + headers

            # 自动猜测字段
            def guess_header(keywords):
                for h in headers:
                    for kw in keywords:
                        if kw in h:
                            return h
                return ""

            self.sample_short_header.set(
                guess_header(["商品简称", "商品名称", "货品标题", "标题", "商品"])
            )
            self.sample_spec_header.set(
                guess_header(["销售规格", "规格", "颜色", "SKU", "sku"])
            )
            self.sample_qty_header.set(
                guess_header(["商品数量", "数量", "件数", "购买数量"])
            )
            self.sample_remark_header.set(
                guess_header(["备注", "买家留言", "商家备注", "订单备注", "留言"])
            )

            messagebox.showinfo(
                "识别完成",
                "已读取样本Excel表头，请检查字段选择是否正确。"
            )

        except Exception as e:
            messagebox.showerror("读取失败", str(e))

    def save_template_from_sample(self):
        """
        根据样本Excel字段选择，生成表头模式导入模板。
        """
        short_col = self.sample_short_header.get().strip()
        spec_col = self.sample_spec_header.get().strip()
        qty_col = self.sample_qty_header.get().strip()
        remark_col = self.sample_remark_header.get().strip()

        if not short_col or not spec_col or not qty_col:
            messagebox.showwarning("提示", "请完整选择：商品简称字段、销售规格字段、数量字段")
            return

        name = self.template_name.get().strip()

        if not name:
            name = f"自定义模板-{short_col}-{spec_col}-{qty_col}"

        self.template_name.set(name)
        self.template_mode.set("表头")
        self.template_short.set(short_col)
        self.template_spec.set(spec_col)
        self.template_qty.set(qty_col)
        self.template_remark.set(remark_col)

        self.save_template()

        messagebox.showinfo(
            "成功",
            f"已保存自定义导入模板：{name}"
        )

    def save_template(self):
        name = self.template_name.get().strip()
        if not name:
            messagebox.showwarning("提示", "模板名称不能为空")
            return

        t = {
            "name": name,
            "mode": self.template_mode.get(),
            "short_name": self.template_short.get().strip(),
            "spec": self.template_spec.get().strip(),
            "qty": self.template_qty.get().strip(),
            "remark": self.template_remark.get().strip(),
            "title_col": self.template_title_col.get().strip(),
            "qty_col": self.template_qty_col.get().strip(),
            "item_sep": self.template_item_sep.get() or ";",
            "spec_split": self.template_spec_split.get() or "，"
        }

        arr = self.system.setdefault("import_templates", [])
        for i, x in enumerate(arr):
            if x.get("name") == name:
                arr[i] = t
                break
        else:
            arr.append(t)

        # 兼容旧数据字段，但不再由后端控制当前使用模板。
        self.system["active_template"] = ""

        self.save_all()
        save_templates_fast(self.system.get("import_templates", []))
        self.refresh_templates()
        messagebox.showinfo("成功", "模板已保存。前端整理时可在模板下拉框中选择使用。")

    def load_template(self, event=None):
        sel = self.template_tree.selection()
        if not sel:
            return

        vals = self.template_tree.item(sel[0]).get("values", [])
        if len(vals) >= 10:
            self.template_name.set(vals[0])
            self.template_mode.set(vals[1])
            self.template_short.set(vals[2])
            self.template_spec.set(vals[3])
            self.template_qty.set(vals[4])
            self.template_remark.set(vals[5])
            self.template_title_col.set(vals[6])
            self.template_qty_col.set(vals[7])
            self.template_item_sep.set(vals[8])
            self.template_spec_split.set(vals[9])

    def delete_template(self):
        sel = self.template_tree.selection()
        if not sel:
            return
        name = self.template_tree.item(sel[0]).get("values", [None])[0]
        if not name:
            return
        if not messagebox.askyesno("确认", f"删除模板：{name}？"):
            return
        self.system["import_templates"] = [t for t in self.system.get("import_templates", []) if t.get("name") != name]
        if self.system.get("active_template") == name:
            self.system["active_template"] = ""
        self.save_all()
        self.refresh_all()

    def preview_excel_headers(self):
        path = filedialog.askopenfilename(title="选择Excel样本", filetypes=[("Excel文件", "*.xlsx *.xls")])
        if not path:
            return
        try:
            df = pd.read_excel(path, nrows=3)
            headers = [str(c).strip() for c in df.columns if str(c).strip()]

            if hasattr(self, "sample_short_combo"):
                self.sample_headers = headers
                self.sample_short_combo["values"] = headers
                self.sample_spec_combo["values"] = headers
                self.sample_qty_combo["values"] = headers

            messagebox.showinfo("检测到表头", "\n".join(map(str, headers[:100])))
        except Exception as e:
            messagebox.showerror("读取失败", str(e))

    def refresh_rules(self):
        for i in self.rule_tree.get_children():
            self.rule_tree.delete(i)
        for r in self.system.get("category_rules", []):
            self.rule_tree.insert("", "end", values=(r.get("category",""), r.get("keyword",""), r.get("field",""), r.get("remove_words","")))

    def save_rule(self):
        cat = normalize_text(self.rule_category.get())
        kw = normalize_text(self.rule_keyword.get())
        field = self.rule_field.get()
        remove = self.rule_remove.get().strip()
        if not cat or not kw:
            messagebox.showwarning("提示", "分类和关键词不能为空")
            return
        rule = {"category": cat, "keyword": kw, "field": field, "remove_words": remove}
        arr = self.system.setdefault("category_rules", [])
        for i, r in enumerate(arr):
            if normalize_text(r.get("category")) == cat and normalize_text(r.get("keyword")) == kw and r.get("field") == field:
                arr[i] = rule
                break
        else:
            arr.append(rule)
        self.save_all()
        self.rule_category.set("")
        self.rule_keyword.set("")
        self.rule_remove.set("")
        self.refresh_all()

    def load_rule(self, event=None):
        sel = self.rule_tree.selection()
        if not sel:
            return
        vals = self.rule_tree.item(sel[0]).get("values", [])
        if len(vals) >= 4:
            self.rule_category.set(vals[0])
            self.rule_keyword.set(vals[1])
            self.rule_field.set(vals[2])
            self.rule_remove.set(vals[3])

    def delete_rule(self):
        sel = self.rule_tree.selection()
        if not sel:
            return
        vals = self.rule_tree.item(sel[0]).get("values", [])
        if len(vals) < 3:
            return
        cat, kw, field = normalize_text(vals[0]), normalize_text(vals[1]), vals[2]
        self.system["category_rules"] = [r for r in self.system.get("category_rules", []) if not (normalize_text(r.get("category")) == cat and normalize_text(r.get("keyword")) == kw and r.get("field") == field)]
        self.save_all()
        self.refresh_all()

    def refresh_stalls(self):
        for i in self.stall_tree.get_children():
            self.stall_tree.delete(i)
        for cat, stall in sorted(self.system.get("stall_map", {}).items()):
            self.stall_tree.insert("", "end", values=(cat, stall))

    def save_stall(self):
        cat = normalize_text(self.stall_category.get())
        stall = self.stall_name.get().strip()
        if not cat or not stall:
            return
        self.system.setdefault("stall_map", {})[cat] = stall
        self.save_all()
        self.stall_category.set("")
        self.stall_name.set("")
        self.refresh_all()

    def load_stall(self, event=None):
        sel = self.stall_tree.selection()
        if not sel:
            return
        vals = self.stall_tree.item(sel[0]).get("values", [])
        if len(vals) >= 2:
            self.stall_category.set(vals[0])
            self.stall_name.set(vals[1])

    def delete_stall(self):
        sel = self.stall_tree.selection()
        if not sel:
            return
        vals = self.stall_tree.item(sel[0]).get("values", [])
        if vals:
            self.system.get("stall_map", {}).pop(normalize_text(vals[0]), None)
            self.save_all()
            self.refresh_all()

    def choose_image(self):
        p = filedialog.askopenfilename(title="选择图片", filetypes=[("图片文件", "*.jpg *.jpeg *.png *.webp *.bmp")])
        if p:
            self.img_path.set(p)

    def save_image(self):
        cat = normalize_text(self.img_category.get())
        spec = normalize_text(self.img_spec.get())
        path = self.img_path.get().strip()
        if not cat or not spec or not os.path.exists(path):
            messagebox.showwarning("提示", "分类、规格、图片不能为空")
            return
        upsert_image_binding(cat, spec, source_path=path)
        save_data(self.data)
        self.img_spec.set("")
        self.img_path.set("")
        self.refresh_all()

    def refresh_images(self):
        if not hasattr(self, "image_tree"):
            return

        for i in self.image_tree.get_children():
            self.image_tree.delete(i)

        keyword = normalize_text(self.image_search_var.get())
        category_filter = self.image_filter_category_var.get()

        shown = 0
        total = 0
        for key, item in iter_image_bindings(category_filter, keyword):
            total += 1
            if "|" in key:
                cat, spec = key.split("|", 1)
            else:
                cat, spec = "", key

            status = "已绑定" if (item.get("image_file") or item.get("file")) else "无图片"
            self.image_tree.insert("", "end", values=(cat, spec, status))
            shown += 1

        if hasattr(self, "image_count_label"):
            stats = image_storage_summary(count_entries=False)
            self.image_count_label.config(
                text=f"当前显示：{shown} 条 / 分类文件：{stats.get('category_files', 0)} 个 / {stats.get('bytes', 0)/1024/1024:.2f} MB"
            )

    def import_images_from_excel(self):
        cat = normalize_text(self.img_category.get())
        if not cat:
            messagebox.showwarning("提示", "请先选择分类")
            return
        excel = filedialog.askopenfilename(title="选择图片规格Excel", filetypes=[("Excel文件", "*.xlsx")])
        if not excel:
            return
        try:
            cell_images = extract_wps_cell_images_from_xlsx(excel)
            if not cell_images:
                messagebox.showwarning("提示", "没有识别到单元格图片")
                return

            wb = openpyxl.load_workbook(excel, data_only=False)
            cnt = 0
            skip = 0
            for _, info in cell_images.items():
                ws = wb[info["sheet"]]
                cell = ws[info["cell"]]
                spec_text = find_spec_right_of_image(ws, cell)
                specs = split_specs_text(spec_text)
                if not specs:
                    skip += 1
                    continue

                for spec in specs:
                    upsert_image_binding(
                        cat,
                        normalize_text(spec),
                        image_bytes=info["bytes"],
                        filename=info.get("filename", "image.png")
                    )
                    cnt += 1

            save_data(self.data)
            self.refresh_all()
            messagebox.showinfo("完成", f"导入 {cnt} 条，跳过 {skip} 条")
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    def reset_image_filter(self):
        self.image_filter_category_var.set("全部分类")
        self.image_search_var.set("")
        self.refresh_images()

    def get_filtered_image_keys(self):
        keyword = normalize_text(self.image_search_var.get())
        category_filter = self.image_filter_category_var.get()
        keys = []

        for key, _ in iter_image_bindings(category_filter, keyword):
            keys.append(key)

        return keys

    def delete_selected_images(self):
        sels = self.image_tree.selection()
        if not sels:
            return
        if not messagebox.askyesno("确认", f"删除选中 {len(sels)} 条图片关系？"):
            return
        for item in sels:
            vals = self.image_tree.item(item).get("values", [])
            if len(vals) >= 2:
                delete_image_binding(vals[0], vals[1])
        save_data(self.data)
        self.refresh_all()

    def delete_filtered_images(self):
        keys = self.get_filtered_image_keys()
        if not keys:
            messagebox.showinfo("提示", "当前筛选结果为空")
            return
        if not messagebox.askyesno("确认删除", f"确认删除当前筛选结果中的 {len(keys)} 条图片关系？"):
            return
        if not messagebox.askyesno("再次确认", "该操作不可撤销，是否继续？"):
            return
        for key in keys:
            if "|" in key:
                cat, spec = key.split("|", 1)
                delete_image_binding(cat, spec)
        save_data(self.data)
        self.refresh_all()

    def clear_all_images(self):
        stats = image_storage_summary(count_entries=True)
        total = stats.get("entries", 0) or 0
        if total and messagebox.askyesno("确认", f"清空全部 {total} 条图片关系？"):
            clear_all_image_categories()
            save_data(self.data)
            self.refresh_all()

    def backup_data(self):
        src = get_data_file()
        if os.path.exists(src):
            dst = src + ".bak"
            shutil.copy2(src, dst)
            messagebox.showinfo("成功", f"已备份：{dst}")

    def compact_data(self):
        try:
            main_size = os.path.getsize(get_data_file()) if os.path.exists(get_data_file()) else 0
            stats = image_storage_summary(count_entries=True)
            self.save_all()
            after = os.path.getsize(get_data_file()) if os.path.exists(get_data_file()) else 0
            messagebox.showinfo(
                "数据大小",
                f"主数据：{after/1024/1024:.2f} MB\n"
                f"保存前主数据：{main_size/1024/1024:.2f} MB\n"
                f"图片分片：{stats.get('category_files', 0)} 个文件\n"
                f"图片关系：{stats.get('entries', 0)} 条\n"
                f"图片数据大小：{stats.get('bytes', 0)/1024/1024:.2f} MB"
            )
        except Exception as e:
            messagebox.showerror("读取失败", str(e))

    def export_config(self):
        self.save_all()
        p = filedialog.asksaveasfilename(title="导出配置包", defaultextension=".enc", initialfile="订单整理系统配置包.enc", filetypes=[("加密配置包", "*.enc")])
        if p:
            shutil.copy2(get_data_file(), p)
            messagebox.showinfo("成功", f"已导出：{p}")

    def import_config(self):
        p = filedialog.askopenfilename(title="导入配置包", filetypes=[("加密配置包", "*.enc")])
        if not p:
            return
        if not messagebox.askyesno("确认", "导入会覆盖当前配置，是否继续？"):
            return
        shutil.copy2(p, get_data_file())
        self.data = load_data(auto_save_on_read=False)
        self.system, self.system_id = get_active_system(self.data)
        self.refresh_all()
        messagebox.showinfo("成功", "配置包已导入")


def main():
    root = tk.Tk()
    set_window_icon(root)
    BackendAdmin(root)
    root.mainloop()


if __name__ == "__main__":
    main()
