def normalize_qty(value):
    try:
        if value is None or str(value).strip() == "":
            return 1
        return int(float(value))
    except Exception:
        return 1


def size_sort_key(size):
    s = str(size).strip()
    try:
        return (0, float(s))
    except Exception:
        return (1, s)


def merge_sizes(values):
    """
    旧兼容函数：仅合并尺码列表。
    """
    vals = [str(v).strip() for v in values if str(v).strip()]
    unique = []
    for v in vals:
        if v not in unique:
            unique.append(v)
    return " ".join(sorted(unique, key=size_sort_key))


def merge_size_quantity(group):
    """
    按商品数量展开尺码。
    例如：41码，数量3，输出为：41 41 41。
    """
    expanded = []

    for _, row in group.iterrows():
        size = str(row.get("尺码", "")).strip()
        if not size:
            continue

        qty = normalize_qty(row.get("数量", 1))
        for _ in range(max(qty, 0)):
            expanded.append(size)

    return " ".join(sorted(expanded, key=size_sort_key))


import os
import re
import sys
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from order_secure_common import (
    load_data, load_templates_fast, get_active_system, get_output_dir, get_temp_dir, get_data_dir,
    normalize_text, make_image_key, base64_to_image_file, safe_filename,
    col_letter_to_index, ImageMatcher, load_image_map_for_categories
)
from order_core import generate_order_file


APP_TITLE = "一键整理订单"
IMAGE_WIDTH_PX = 140
IMAGE_HEIGHT_PX = 120


def resource_path(filename):
    if getattr(sys, "frozen", False):
        base_dir = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(sys.executable)))
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, filename)


def set_window_icon(root):
    try:
        ico = resource_path("icon_frontend.ico")
        if os.path.exists(ico):
            root.iconbitmap(default=ico)
    except Exception:
        pass
    try:
        png = resource_path("icon_frontend.png")
        if os.path.exists(png):
            img = tk.PhotoImage(file=png)
            root.iconphoto(True, img)
            root._icon_ref = img
    except Exception:
        pass


def split_items(value, sep):
    if value is None:
        return []
    return [x.strip() for x in str(value).split(sep) if str(x).strip()]


def split_spec_to_color_size(spec, split_char):
    spec = str(spec).strip()
    for sp in [split_char, "，", ",", " "]:
        if sp and sp in spec:
            parts = spec.rsplit(sp, 1)
            if len(parts) == 2:
                return parts[0].strip(), parts[1].strip()
    return spec, "未知"


def read_by_template(file_path, template):
    df = pd.read_excel(file_path)
    mode = template.get("mode", "表头")
    item_sep = template.get("item_sep", ";") or ";"
    spec_split = template.get("spec_split", "，") or "，"
    rows = []

    if mode == "表头":
        short_col = template.get("short_name", "")
        spec_col = template.get("spec", "")
        qty_col = template.get("qty", "")

        if not {short_col, spec_col, qty_col}.issubset(set(df.columns)):
            raise ValueError(f"文件缺少模板指定表头：{os.path.basename(file_path)}")

        temp = df[[short_col, spec_col, qty_col]].copy()
        temp = temp.dropna(subset=[short_col, spec_col])

        for _, row in temp.iterrows():
            names = split_items(row[short_col], item_sep)
            specs = split_items(row[spec_col], item_sep)
            qtys = split_items(row[qty_col], item_sep)
            max_len = max(len(names), len(specs), len(qtys), 1)

            for i in range(max_len):
                name = names[i] if i < len(names) else (names[-1] if names else "")
                spec = specs[i] if i < len(specs) else (specs[-1] if specs else "")
                qty = qtys[i] if i < len(qtys) else (qtys[-1] if qtys else "1")
                color, size = split_spec_to_color_size(spec, spec_split)
                title = f"{name} 颜色: {color} 尺码: {size}"
                rows.append({"商品简称": name, "货品标题": title, "数量": qty, "来源文件": os.path.basename(file_path)})

    else:
        title_idx = col_letter_to_index(template.get("title_col", "S"))
        qty_idx = col_letter_to_index(template.get("qty_col", "V"))
        if df.shape[1] <= max(title_idx, qty_idx):
            raise ValueError(f"文件列数不足：{os.path.basename(file_path)}")

        temp = df.iloc[:, [title_idx, qty_idx]].copy()
        temp.columns = ["货品标题", "数量"]
        temp = temp.dropna(subset=["货品标题"])

        def guess_short_name(title):
            m = re.search(r"^(.*?)\s*颜色[:：]", str(title))
            return m.group(1).strip() if m else str(title).strip()

        for _, row in temp.iterrows():
            rows.append({"商品简称": guess_short_name(row["货品标题"]), "货品标题": row["货品标题"], "数量": row["数量"], "来源文件": os.path.basename(file_path)})

    out = pd.DataFrame(rows)
    out["数量"] = pd.to_numeric(out["数量"], errors="coerce").fillna(1).astype(int)
    return out


def extract_size(title):
    m = re.search(r"尺码[:：]\s*([0-9.]+)", str(title))
    return m.group(1) if m else "未知"


def extract_raw_spec(title):
    m = re.search(r"颜色[:：]\s*(.*?)\s*尺码", str(title))
    return normalize_text(m.group(1)) if m else "未知"


def size_sort_key(size):
    try:
        return float(size)
    except Exception:
        return 999


def merge_sizes(series):
    sizes = [str(x).strip() for x in series if str(x).strip() and str(x).strip() != "未知"]
    return " ".join(sorted(sizes, key=size_sort_key))


def rule_score(rule, short_name, title, raw_spec):
    kw = normalize_text(rule.get("keyword", ""))
    if not kw:
        return -1
    field = rule.get("field", "全部")

    if field == "商品简称":
        targets = [(normalize_text(short_name), 1000)]
    elif field == "销售规格":
        targets = [(normalize_text(raw_spec), 700)]
    elif field == "货品标题":
        targets = [(normalize_text(title), 500)]
    else:
        targets = [(normalize_text(short_name), 1000), (normalize_text(raw_spec), 700), (normalize_text(title), 500)]

    best = -1
    for text, weight in targets:
        if kw in text:
            best = max(best, weight + len(kw))
    return best


def detect_category(short_name, title, raw_spec, rules):
    category = "未分类"
    best = -1
    for r in rules:
        score = rule_score(r, short_name, title, raw_spec)
        if score > best:
            category = r.get("category", "未分类")
            best = score
    return category


def clean_spec(raw_spec, category, rules):
    spec = normalize_text(raw_spec)
    words = []
    for r in rules:
        if normalize_text(r.get("category", "")) == normalize_text(category):
            remove_words = str(r.get("remove_words", "")).strip()
            if remove_words:
                for w in remove_words.replace("，", ",").split(","):
                    w = normalize_text(w)
                    if w and w not in words:
                        words.append(w)
    for w in words:
        spec = spec.replace(w, "")
    return spec if spec else raw_spec


def get_stall(category, stall_map):
    return stall_map.get(normalize_text(category), "未设置档口")


def safe_sheet_name(name):
    name = re.sub(r'[\\/:*?\[\]]', "_", str(name).strip())
    return name[:31] if name else "未设置档口"


def prepare_image_for_excel(image_b64, category, spec):
    temp_dir = get_temp_dir()
    raw = os.path.join(temp_dir, f"{safe_filename(category)}_{safe_filename(spec)}_raw")
    png = os.path.join(temp_dir, f"{safe_filename(category)}_{safe_filename(spec)}.png")
    base64_to_image_file(image_b64, raw)
    return prepare_image_file_for_excel(raw, category, spec)


def prepare_image_file_for_excel(source_path, category, spec):
    temp_dir = get_temp_dir()
    png = os.path.join(temp_dir, f"{safe_filename(category)}_{safe_filename(spec)}.png")

    img = PILImage.open(source_path)
    if img.mode in ("RGBA", "LA"):
        bg = PILImage.new("RGBA", img.size, (255, 255, 255, 255))
        bg.paste(img, mask=img.getchannel("A"))
        img = bg.convert("RGB")
    else:
        img = img.convert("RGB")

    img.thumbnail((IMAGE_WIDTH_PX, IMAGE_HEIGHT_PX), PILImage.LANCZOS)
    canvas = PILImage.new("RGB", (IMAGE_WIDTH_PX, IMAGE_HEIGHT_PX), (255, 255, 255))
    canvas.paste(img, ((IMAGE_WIDTH_PX - img.width) // 2, (IMAGE_HEIGHT_PX - img.height) // 2))
    canvas.save(png)
    return png


def prepare_image_item_for_excel(item, category, spec):
    if not item:
        return ""

    image_file = item.get("image_file") or item.get("file") or ""
    if image_file:
        source = image_file
        if not os.path.isabs(source):
            source = os.path.join(get_data_dir(), image_file)
        if os.path.exists(source):
            return prepare_image_file_for_excel(source, category, spec)

    image_b64 = item.get("image_base64")
    if image_b64:
        return prepare_image_for_excel(image_b64, category, spec)

    return ""


def style_and_images(output_file, image_map):
    matcher = ImageMatcher(image_map)
    wb = load_workbook(output_file)

    for ws in wb.worksheets:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = {normalize_text(ws.cell(row=1, column=col).value): col for col in range(1, ws.max_column + 1)}
        category_col = headers.get("分类", 1)
        spec_col = headers.get("规格", 2)
        image_col = headers.get("图片", 3)
        image_col_letter = ws.cell(row=1, column=image_col).column_letter

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for col in range(1, ws.max_column + 1):
            col_letter = ws.cell(row=1, column=col).column_letter
            header = normalize_text(ws.cell(row=1, column=col).value)
            if header == "图片":
                ws.column_dimensions[col_letter].width = 20
            elif header == "尺码":
                ws.column_dimensions[col_letter].width = 42
            elif header == "规格":
                ws.column_dimensions[col_letter].width = 28
            else:
                ws.column_dimensions[col_letter].width = 16

        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = IMAGE_HEIGHT_PX * 0.75
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            category = normalize_text(ws.cell(row=row, column=category_col).value)
            spec = normalize_text(ws.cell(row=row, column=spec_col).value)
            item = image_map.get(make_image_key(category, spec))

            if item and item.get("image_base64"):
                try:
                    png_path = prepare_image_for_excel(item["image_base64"], category, spec)
                    img = XLImage(png_path)
                    img.width = IMAGE_WIDTH_PX
                    img.height = IMAGE_HEIGHT_PX
                    ws.add_image(img, f"{image_col_letter}{row}")
                    ws.cell(row=row, column=image_col).value = ""
                except Exception:
                    ws.cell(row=row, column=image_col).value = ""

    wb.save(output_file)


def generate(files, output_mode, template_name):
    data = load_data(auto_save_on_read=False)
    system, system_id = get_active_system(data)

    return generate_order_file(files, system, output_mode, template_name)


class FrontendApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("820x520")
        self.root.minsize(720, 460)

        self.files = []
        self.file_text = tk.StringVar(value="请选择一个或多个订单Excel")
        self.output_mode = tk.StringVar(value="按档口分Sheet")
        self.template_name = tk.StringVar(value="")
        self.template_combo = None

        self.build_ui()
        self.refresh_templates()

    def build_ui(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        top = ttk.LabelFrame(self.root, text="订单整理")
        top.grid(row=0, column=0, sticky="ew", padx=16, pady=12)
        top.columnconfigure(0, weight=1)

        ttk.Entry(top, textvariable=self.file_text).grid(row=0, column=0, sticky="ew", padx=8, pady=8)
        ttk.Button(top, text="选择多个Excel", command=self.choose_files).grid(row=0, column=1, padx=8)
        tk.Button(top, text="生成整理文档", command=self.generate_file, bg="#1F4E78", fg="white", width=16).grid(row=0, column=2, padx=8)

        template_frame = ttk.Frame(top)
        template_frame.grid(row=1, column=0, columnspan=3, sticky="w", padx=8, pady=(4, 8))
        ttk.Label(template_frame, text="导入模板：").pack(side="left")
        self.template_combo = ttk.Combobox(
            template_frame,
            textvariable=self.template_name,
            values=[],
            state="readonly",
            width=36
        )
        self.template_combo.pack(side="left", padx=8)
        ttk.Button(template_frame, text="刷新模板", command=self.refresh_templates).pack(side="left", padx=4)

        mode_frame = ttk.Frame(top)
        mode_frame.grid(row=2, column=0, columnspan=3, sticky="w", padx=8, pady=8)
        ttk.Label(mode_frame, text="输出方式：").pack(side="left")
        for mode in ["合并一个Sheet", "按档口分Sheet", "按档口分文档"]:
            ttk.Radiobutton(mode_frame, text=mode, variable=self.output_mode, value=mode).pack(side="left", padx=8)

        self.listbox = tk.Listbox(self.root)
        self.listbox.grid(row=1, column=0, sticky="nsew", padx=16, pady=8)

        bottom = ttk.Frame(self.root)
        bottom.grid(row=2, column=0, sticky="ew", padx=16, pady=12)
        ttk.Button(bottom, text="清空文件列表", command=self.clear_files).pack(side="left", padx=5)
        ttk.Button(bottom, text="打开输出目录", command=lambda: os.startfile(get_output_dir())).pack(side="left", padx=5)

    def refresh_templates(self):
        try:
            # 只刷新模板列表，不读取订单、不扫描图片、不整库重写。
            template_items = load_templates_fast()
            templates = [t.get("name", "") for t in template_items if t.get("name", "")]

            if self.template_combo is not None:
                self.template_combo["values"] = templates

            current = self.template_name.get()
            if templates and current not in templates:
                self.template_name.set(templates[0])
            elif not templates:
                self.template_name.set("")
        except Exception:
            if self.template_combo is not None:
                self.template_combo["values"] = []
            self.template_name.set("")

    def choose_files(self):
        files = filedialog.askopenfilenames(title="选择订单Excel", filetypes=[("Excel文件", "*.xlsx *.xls")])
        for f in files:
            if f not in self.files:
                self.files.append(f)
        self.refresh_files()

    def refresh_files(self):
        self.listbox.delete(0, tk.END)
        for f in self.files:
            self.listbox.insert(tk.END, f)
        self.file_text.set(f"已选择 {len(self.files)} 个Excel文件" if self.files else "请选择一个或多个订单Excel")

    def clear_files(self):
        self.files = []
        self.refresh_files()

    def generate_file(self):
        if not self.files:
            messagebox.showwarning("提示", "请先选择订单Excel")
            return
        template_name = self.template_name.get().strip()
        if not template_name:
            messagebox.showwarning("提示", "请先选择导入模板")
            return

        try:
            out = generate(self.files, self.output_mode.get(), template_name)
            messagebox.showinfo("完成", f"整理文档已生成：\n{out}")
        except Exception as e:
            messagebox.showerror("生成失败", str(e))


def main():
    root = tk.Tk()
    set_window_icon(root)
    FrontendApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
