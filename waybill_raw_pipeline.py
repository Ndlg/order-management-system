from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from waybill_raw_contract import (
    AUXILIARY_WAYBILL_FIELDS,
    PROCESSED_WAYBILL_FIELDS,
    RAW_PIPELINE_INTERNAL_FIELDS,
    RAW_WAYBILL_MODE,
    RAW_WAYBILL_TEXT_COLUMN,
    RAW_WAYBILL_TRACKING_FIELDS,
    is_raw_waybill_template_name,
    raw_waybill_text_column,
)
from waybill_text_parser import normalize_raw_text, parse_waybill_raw_text


NO_PRINT_TEXT_STATUS = "无打印信息"
UNPARSED_STATUS = "无法解析"


def is_waybill_raw_template(template: dict, columns) -> bool:
    headers = {str(col).strip() for col in columns}
    raw_col = raw_waybill_text_column(template)
    if not raw_col or raw_col not in headers:
        return False

    if is_raw_waybill_template_name(template.get("name")):
        return True
    if str(template.get("mode") or "").strip() == RAW_WAYBILL_MODE:
        return True

    # Compatibility with early test builds that used one header as three fields.
    return (
        raw_col == RAW_WAYBILL_TEXT_COLUMN
        and raw_col == str(template.get("spec") or "").strip() == str(template.get("qty") or "").strip()
    )


def parse_raw_waybill_dataframe(
    df: pd.DataFrame,
    file_path: str,
    template: dict | None = None,
    rule_config: object | None = None,
) -> list[dict]:
    raw_col = raw_waybill_text_column(template)
    if raw_col not in df.columns:
        raise ValueError(f"文件缺少打印信息表头：{os.path.basename(file_path)}")

    rows: list[dict] = []
    temp = df.copy()
    for _, row in temp.iterrows():
        raw_text = normalize_raw_text(row.get(raw_col, ""))
        if not raw_text:
            if row_has_tracking(row):
                rows.append(_raw_parse_row("", {}, NO_PRINT_TEXT_STATUS, row))
            continue

        parsed_rows = parse_waybill_raw_text(raw_text, "", rule_config)
        if not parsed_rows:
            rows.append(_raw_parse_row(raw_text, {}, UNPARSED_STATUS, row))
            continue

        for parsed in parsed_rows:
            rows.append(_raw_parse_row(raw_text, parsed, "已解析", row))

    return rows


def raw_text_from_record(record: dict) -> str:
    return normalize_raw_text(
        record.get("打印信息")
        or record.get("print_text_raw")
        or record.get("print_text")
        or ""
    )


def parse_raw_waybill_records(records: list[dict], rule_config: object | None = None) -> list[dict]:
    rows: list[dict] = []
    for record in records:
        raw_text = raw_text_from_record(record)
        if not raw_text:
            rows.append(_raw_parse_row("", {}, NO_PRINT_TEXT_STATUS, record))
            continue

        parsed_rows = parse_waybill_raw_text(raw_text, "", rule_config)
        if not parsed_rows:
            rows.append(_raw_parse_row(raw_text, {}, UNPARSED_STATUS, record))
            continue

        for parsed in parsed_rows:
            rows.append(_raw_parse_row(raw_text, parsed, "已解析", record))
    return rows


def tracking_values(source: object | None) -> dict:
    def get_value(key: str, default: str = "") -> str:
        if source is None:
            return default
        if isinstance(source, dict):
            value = source.get(key, default)
        else:
            try:
                value = source.get(key, default)
            except AttributeError:
                value = default
        if pd.isna(value):
            return ""
        return str(value).strip()

    source_machine = (
        get_value("来源机器")
        or get_value("machine_label")
        or get_value("machine_name")
    )
    return {
        "任务ID": get_value("任务ID") or get_value("task_id"),
        "文档ID": get_value("文档ID") or get_value("document_id"),
        "任务时间": get_value("任务时间") or get_value("task_time"),
        "采集端ID": get_value("采集端ID") or get_value("source_client_id"),
        "来源机器": source_machine,
        "来源序号": get_value("来源序号") or get_value("source_record_index") or get_value("record_index"),
    }


def row_has_tracking(row: object) -> bool:
    if row is None:
        return False
    for field in RAW_WAYBILL_TRACKING_FIELDS:
        try:
            value = row.get(field, "")
        except AttributeError:
            value = ""
        if pd.notna(value) and str(value).strip():
            return True
    return False


def _raw_parse_row(raw_text: str, parsed: dict, status: str, source: object | None = None) -> dict:
    row = {field: parsed.get(field, "") for field in AUXILIARY_WAYBILL_FIELDS + PROCESSED_WAYBILL_FIELDS}
    tracking = tracking_values(source)
    row.update({field: tracking.get(field, "") for field in RAW_WAYBILL_TRACKING_FIELDS})
    row["原始打印信息"] = raw_text
    row["解析状态"] = status
    return row


def processed_waybill_dataframe(rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(rows, columns=RAW_PIPELINE_INTERNAL_FIELDS)


def write_processed_waybill_xlsx(rows: list[dict], path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "面单解析结果"
    for col, header in enumerate(RAW_PIPELINE_INTERNAL_FIELDS, 1):
        ws.cell(row=1, column=col, value=header)
    for row_index, row in enumerate(rows, 2):
        for col, header in enumerate(RAW_PIPELINE_INTERNAL_FIELDS, 1):
            ws.cell(row=row_index, column=col, value=row.get(header, ""))

    tmp = path.with_suffix(".tmp.xlsx")
    try:
        wb.save(tmp)
        os.replace(tmp, path)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
    return path
